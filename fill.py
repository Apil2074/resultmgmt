import openpyxl
import random
import math
import re

# File paths
file_path = "ME_All_Classes_FIRST_TERMINAL_EXAMINATION.xlsx"
out_path = "ME_All_Classes_FILLED.xlsx"

# Load the workbook
wb = openpyxl.load_workbook(file_path)

# --- NEW CONFIGURATION ---
TARGET_PASS_RATE = 0.85      # 85% of the class will pass every subject
PASSING_PERCENTAGE = 0.40    # 40% of full marks is required to pass
TOTAL_WORKING_DAYS = 65      # Total attendance days for the term
# -------------------------

def get_class_level(sheet_name):
    sheet_name = sheet_name.lower()
    if any(x in sheet_name for x in ['nursery', 'ukg', 'lkg', 'kg']):
        return 8 
    
    match = re.search(r'\d+', sheet_name)
    if match:
        return int(match.group())
    return None

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    level = get_class_level(sheet_name)
    
    if level is None or level > 10:
        continue
        
    header_row, fm_row = None, None
    for r in range(1, 20):
        row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, 5)]
        if any('student name' in v for v in row_vals):
            header_row = r
        if any('full marks' in v for v in row_vals):
            fm_row = r
            
    if not header_row or not fm_row:
        continue
        
    theory_cols, practical_cols = [], []
    present_col = None
    
    for c in range(1, ws.max_column + 1):
        h_val = ws.cell(row=header_row, column=c).value
        if h_val and isinstance(h_val, str):
            h_lower = h_val.lower()
            if 'theory' in h_lower:
                fm_val = ws.cell(row=fm_row, column=c).value
                if fm_val:
                    try:
                        theory_cols.append((c, float(fm_val)))
                    except ValueError:
                        pass
            elif 'internal' in h_lower or 'practical' in h_lower:
                fm_val = ws.cell(row=fm_row, column=c).value
                if fm_val:
                    try:
                        practical_cols.append((c, float(fm_val)))
                    except ValueError:
                        pass
            elif 'present days' in h_lower or 'attendance' in h_lower:
                present_col = c
                # Fill the total attendance days in the Full Marks row
                ws.cell(row=fm_row, column=c).value = TOTAL_WORKING_DAYS
                
    start_row = fm_row + 1
    for r in range(start_row, ws.max_row + 1):
        if not ws.cell(row=r, column=2).value:
            continue
            
        is_guaranteed_pass = random.random() < TARGET_PASS_RATE
        
        for c, th_fm in theory_cols:
            if is_guaranteed_pass:
                min_val = math.ceil(PASSING_PERCENTAGE * th_fm) 
            else:
                min_val = math.ceil(0.25 * th_fm) 
                
            max_val = math.floor(0.95 * th_fm)
            if min_val <= max_val:
                ws.cell(row=r, column=c).value = random.randint(min_val, max_val)
            else:
                ws.cell(row=r, column=c).value = max_val
            
        for c, pr_fm in practical_cols:
            if is_guaranteed_pass:
                min_val = math.ceil(PASSING_PERCENTAGE * pr_fm)
            else:
                min_val = math.ceil(0.25 * pr_fm)
                
            max_val = math.floor(0.95 * pr_fm)
            if min_val <= max_val:
                ws.cell(row=r, column=c).value = random.randint(min_val, max_val)
            else:
                ws.cell(row=r, column=c).value = max_val
            
        if present_col:
            # Generate random attendance between a set minimum (e.g., 45) and the total working days
            ws.cell(row=r, column=present_col).value = random.randint(45, TOTAL_WORKING_DAYS)

wb.save(out_path)
print(f"Successfully processed and saved to {out_path}")