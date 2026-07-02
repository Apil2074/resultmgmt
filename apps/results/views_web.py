"""
Results App — Web views (ledger, marksheet, processing)
"""
import datetime
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django_ratelimit.decorators import ratelimit

from apps.exams.models import ExamClass
from apps.subjects.models import Subject
from apps.students.models import Student
from apps.marks.models import MarkEntry
from .models import StudentResult, SubjectResult
from .services import ResultProcessingService

logger = logging.getLogger(__name__)

def format_mark(val):
    if val is None:
        return '—'
    from decimal import Decimal
    try:
        dec = Decimal(str(val))
        if dec == dec.to_integral_value():
            return str(int(dec))
        return str(dec.normalize())
    except Exception:
        return str(val)





@login_required
def grade_ledger(request, exam_id, class_id):
    """Grade Ledger — horizontal table showing all subject marks."""
    school = request.user.school
    from apps.exams.models import Exam
    from apps.classes.models import Class
    from apps.subjects.models import Subject
    from apps.marks.models import MarkEntry
    from apps.results.models import SubjectResult, StudentResult

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    cls = get_object_or_404(Class, pk=class_id, school=school)

    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=cls)
    if not exam_class.is_locked:
        messages.warning(request, f"Results for {cls.name} have not been processed and locked yet.")
        return redirect('grade_ledger_select')

    subjects = Subject.objects.filter(
        class_obj=cls, school=school
    ).select_related('marking_structure').order_by('order')

    # Get all student results
    student_results = StudentResult.objects.filter(
        exam=exam, student__class_obj=cls, school=school
    ).select_related('student').order_by('class_rank', 'student__roll_number')

    # Build subject-keyed mark map
    mark_map = {}
    for me in MarkEntry.objects.filter(
        exam=exam, school=school, student__class_obj=cls
    ).select_related('subject_result'):
        key = (me.student_id, me.subject_id)
        mark_map[key] = me

    # Build student specific score matrices
    for sr in student_results:
        sr.subject_scores = []
        total_present = 0
        total_days = 0
        has_attendance = False
        
        for subject in subjects:
            me = mark_map.get((sr.student_id, subject.id))
            score_info = {
                'subject': subject,
                'has_theory': subject.marking_structure.has_theory if hasattr(subject, 'marking_structure') else True,
                'theory_full_marks': subject.marking_structure.theory_full_marks if hasattr(subject, 'marking_structure') else 100,
                'theory_obtained': '—',
                'theory_grade': '—',
                'has_internal': subject.marking_structure.has_internal if hasattr(subject, 'marking_structure') else False,
                'internal_full_marks': subject.marking_structure.internal_full_marks if hasattr(subject, 'marking_structure') and subject.marking_structure.has_internal else 0,
                'internal_obtained': '—',
                'internal_grade': '—',
            }
            if me:
                if me.present_days is not None:
                    total_present += me.present_days
                    has_attendance = True
                if me.total_days:
                    total_days += me.total_days
                
                if me.special_value:
                    val = me.get_special_value_display()
                    score_info['theory_obtained'] = val
                    score_info['theory_grade'] = val
                    score_info['theory_grade_point'] = val
                    if score_info['has_internal']:
                        score_info['internal_obtained'] = val
                        score_info['internal_grade'] = val
                        score_info['internal_grade_point'] = val
                    score_info['grade_point'] = val
                    score_info['grade'] = val
                else:
                    if me.theory_obtained is not None:
                        score_info['theory_obtained'] = float(me.theory_obtained)
                    if hasattr(me, 'subject_result') and me.subject_result:
                        score_info['theory_grade'] = me.subject_result.theory_grade or '—'
                        score_info['theory_grade_point'] = me.subject_result.theory_grade_point or 0.0
                    else:
                        score_info['theory_grade'] = '—'
                        score_info['theory_grade_point'] = '—'
                    
                    if score_info['has_internal']:
                        if me.internal_obtained is not None:
                            score_info['internal_obtained'] = float(me.internal_obtained)
                        if hasattr(me, 'subject_result') and me.subject_result:
                            score_info['internal_grade'] = me.subject_result.internal_grade or '—'
                            score_info['internal_grade_point'] = me.subject_result.internal_grade_point or 0.0
                        else:
                            score_info['internal_grade'] = '—'
                            score_info['internal_grade_point'] = '—'
                    
                    if hasattr(me, 'subject_result') and me.subject_result:
                        score_info['grade_point'] = me.subject_result.grade_point or 0.0
                        score_info['grade'] = me.subject_result.grade or '—'
                    else:
                        score_info['grade_point'] = '—'
                        score_info['grade'] = '—'
            sr.subject_scores.append(score_info)
            
        if has_attendance and total_days > 0:
            sr.attendance_pct = round((total_present / total_days) * 100, 1)
        else:
            sr.attendance_pct = None

    return render(request, 'results/ledger.html', {
        'exam': exam,
        'class_obj': cls,
        'subjects': subjects,
        'student_results': student_results,
    })


@login_required
def marksheet(request, exam_id, student_id):
    """Individual student marksheet."""
    school = request.user.school
    from apps.exams.models import Exam
    from apps.students.models import Student
    from apps.results.models import StudentResult
    from apps.marks.models import MarkEntry

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    student = get_object_or_404(Student, pk=student_id, school=school)

    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=student.class_obj)
    if not exam_class.is_locked:
        messages.warning(request, f"Results for {student.class_obj.name} have not been processed and locked yet.")
        return redirect('marksheet_select')

    try:
        result = StudentResult.objects.get(exam=exam, student=student, school=school)
    except StudentResult.DoesNotExist:
        result = None

    mark_entries = MarkEntry.objects.filter(
        exam=exam, student=student, school=school
    ).select_related(
        'subject', 'subject__marking_structure', 'subject_result'
    ).order_by('subject__order')

    grade_remarks = {
        'A+': 'Outstanding',
        'A': 'Excellent',
        'B+': 'Very Good',
        'B': 'Good',
        'C+': 'Satisfactory',
        'C': 'Acceptable',
        'D+': 'Basic',
        'D': 'Basic',
        'E': 'Insufficient',
        'NG': 'Not Graded',
    }
    for me in mark_entries:
        sr = getattr(me, 'subject_result', None)
        if sr:
            sr.theory_remark = grade_remarks.get(sr.theory_grade, '—') if sr.theory_grade else '—'
            sr.internal_remark = grade_remarks.get(sr.internal_grade, '—') if sr.internal_grade else '—'
            sr.overall_remark = grade_remarks.get(sr.grade, '—') if sr.grade else '—'

    present_days = '—'
    total_days = '—'
    for me in mark_entries:
        if me.present_days is not None:
            present_days = me.present_days
        if me.total_days:
            total_days = me.total_days
        if present_days != '—' and total_days != '—':
            break

    # Navigation — prev/next student
    all_students = list(
        student.class_obj.students.filter(is_active=True).order_by('roll_number')
        .values_list('id', flat=True)
    )
    current_idx = all_students.index(student_id) if student_id in all_students else 0
    prev_student_id = all_students[current_idx - 1] if current_idx > 0 else None
    next_student_id = all_students[current_idx + 1] if current_idx < len(all_students) - 1 else None

    credit_mark_entries = [me for me in mark_entries if me.subject.subject_type != 'NON_CREDIT']
    non_credit_mark_entries = [me for me in mark_entries if me.subject.subject_type == 'NON_CREDIT']

    return render(request, 'results/marksheet.html', {
        'exam': exam,
        'student': student,
        'result': result,
        'credit_mark_entries': credit_mark_entries,
        'non_credit_mark_entries': non_credit_mark_entries,
        'prev_student_id': prev_student_id,
        'next_student_id': next_student_id,
        'present_days': present_days,
        'total_days': total_days,
    })


@login_required
def marksheet_pdf(request, exam_id, student_id):
    """Generate PDF marksheet for a student."""
    from apps.reports.pdf_generators import MarksheetPDFGenerator
    school = request.user.school
    from apps.exams.models import Exam
    from apps.students.models import Student
    from apps.results.models import StudentResult
    from apps.marks.models import MarkEntry

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    student = get_object_or_404(Student, pk=student_id, school=school)

    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=student.class_obj)
    if not exam_class.is_locked:
        messages.warning(request, f"Results for {student.class_obj.name} have not been processed and locked yet.")
        return redirect('marksheet_select')

    try:
        result = StudentResult.objects.get(exam=exam, student=student)
    except StudentResult.DoesNotExist:
        messages.error(request, 'Result not yet processed.')
        return redirect('marksheet', exam_id=exam_id, student_id=student_id)

    mark_entries = MarkEntry.objects.filter(
        exam=exam, student=student, school=school
    ).select_related('subject', 'subject__marking_structure', 'subject_result')

    template_version = request.GET.get('template_version', 'default')
    if template_version == 'neb11':
        from apps.reports.pdf_generators import NEB11MarksheetPDFGenerator
        generator = NEB11MarksheetPDFGenerator(school, exam, student, result, mark_entries)
    else:
        from apps.reports.pdf_generators import MarksheetPDFGenerator
        generator = MarksheetPDFGenerator(school, exam, student, result, mark_entries)
        
    pdf_bytes = generator.generate()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="{student.name}_marksheet.pdf"'
    )
    return response


@login_required
def class_marksheets_pdf(request, exam_id, class_id):
    """Generate merged PDF marksheet for all students in a class."""
    from apps.reports.pdf_generators import ClassMarksheetsPDFGenerator
    school = request.user.school
    from apps.exams.models import Exam
    from apps.classes.models import Class
    from apps.results.models import StudentResult
    from apps.marks.models import MarkEntry

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    cls = get_object_or_404(Class, pk=class_id, school=school)

    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=cls)
    if not exam_class.is_locked:
        messages.warning(request, f"Results for {cls.name} have not been processed and locked yet.")
        return redirect('marksheet_select')

    student_results = StudentResult.objects.filter(
        exam=exam, student__class_obj=cls, school=school
    ).select_related('student').order_by('class_rank', 'student__roll_number')

    if not student_results.exists():
        messages.error(request, 'No student results processed yet for this class.')
        return redirect('marksheet_select')

    # Get all mark entries for this exam and class
    mark_entries = MarkEntry.objects.filter(
        exam=exam, school=school, student__class_obj=cls
    ).select_related('subject', 'subject__marking_structure', 'subject_result')

    # Group mark entries by student_id
    from collections import defaultdict
    student_mark_map = defaultdict(list)
    for me in mark_entries:
        student_mark_map[me.student_id].append(me)

    template_version = request.GET.get('template_version', 'default')
    if template_version == 'neb11':
        from apps.reports.pdf_generators import NEB11ClassMarksheetsPDFGenerator
        generator = NEB11ClassMarksheetsPDFGenerator(school, exam, cls, student_results, student_mark_map)
    else:
        from apps.reports.pdf_generators import ClassMarksheetsPDFGenerator
        generator = ClassMarksheetsPDFGenerator(school, exam, cls, student_results, student_mark_map)
        
    pdf_bytes = generator.generate()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="{cls.full_name}_merged_marksheets.pdf"'
    )
    return response


@login_required
def ledger_pdf(request, exam_id, class_id):
    """Generate PDF grade ledger."""
    from apps.reports.pdf_generators import LedgerPDFGenerator
    school = request.user.school
    from apps.exams.models import Exam
    from apps.classes.models import Class
    from apps.subjects.models import Subject
    from apps.marks.models import MarkEntry
    from apps.results.models import StudentResult

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    cls = get_object_or_404(Class, pk=class_id, school=school)

    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=cls)
    if not exam_class.is_locked:
        messages.warning(request, f"Results for {cls.name} have not been processed and locked yet.")
        return redirect('grade_ledger_select')

    subjects = Subject.objects.filter(class_obj=cls, school=school).select_related(
        'marking_structure').order_by('order')
    student_results_qs = StudentResult.objects.filter(
        exam=exam, student__class_obj=cls
    ).select_related('student')
    
    def student_roll_sort_key(sr):
        try:
            return int(sr.student.roll_number)
        except Exception:
            return sr.student.roll_number
            
    student_results = list(student_results_qs)
    student_results.sort(key=student_roll_sort_key)
    mark_map = {}
    for me in MarkEntry.objects.filter(exam=exam, school=school, student__class_obj=cls
                                        ).select_related('subject_result'):
        mark_map[(me.student_id, me.subject_id)] = me

    generator = LedgerPDFGenerator(school, exam, cls, subjects, student_results, mark_map)
    pdf_bytes = generator.generate()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="{cls.full_name}_ledger.pdf"'
    )
    return response


@login_required
def ledger_excel(request, exam_id, class_id):
    """Generate Excel grade ledger."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    school = request.user.school
    from apps.exams.models import Exam
    from apps.classes.models import Class
    from apps.subjects.models import Subject
    from apps.marks.models import MarkEntry
    from apps.results.models import StudentResult

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    cls = get_object_or_404(Class, pk=class_id, school=school)

    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=cls)
    if not exam_class.is_locked:
        messages.warning(request, f"Results for {cls.name} have not been processed and locked yet.")
        return redirect('grade_ledger_select')

    subjects = list(Subject.objects.filter(class_obj=cls, school=school).select_related(
        'marking_structure').order_by('order'))
        
    student_results_qs = StudentResult.objects.filter(
        exam=exam, student__class_obj=cls
    ).select_related('student')
    
    def student_roll_sort_key(sr):
        try:
            return int(sr.student.roll_number)
        except Exception:
            return sr.student.roll_number
            
    student_results = list(student_results_qs)
    student_results.sort(key=student_roll_sort_key)
    
    mark_map = {}
    for me in MarkEntry.objects.filter(exam=exam, school=school, student__class_obj=cls
                                        ).select_related('subject_result'):
        mark_map[(me.student_id, me.subject_id)] = me

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ledger - {cls.name}"[:31]

    # Header Row
    headers = ["SN", "Symbol No", "Reg No", "Student Name", "Date of Birth"]
    
    for subject in subjects:
        if subject.marking_structure.has_theory:
            headers.append(f"{subject.name} - Th")
        if subject.marking_structure.has_internal:
            headers.append(f"{subject.name} - In")
        headers.extend([
            f"{subject.name} - Th GP",
            f"{subject.name} - In GP",
            f"{subject.name} - Grade"
        ])
    
    headers.extend(["Total Credit", "GPA", "Final Grade", "Rank", "Result"])
    
    ws.append(headers)
    
    # Style Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Add Data
    for idx, sr in enumerate(student_results, 1):
        row = [
            idx,
            sr.student.symbol_number or "",
            sr.student.registration_number or "",
            sr.student.name,
            sr.student.dob_full or ""
        ]
        
        for subject in subjects:
            me = mark_map.get((sr.student_id, subject.id))
            if not me or not me.subject_result:
                # Add empty columns based on subject structure
                cols = 3
                if subject.marking_structure.has_theory: cols += 1
                if subject.marking_structure.has_internal: cols += 1
                row.extend([""] * cols)
                continue
                
            sr_subj = me.subject_result
            if subject.marking_structure.has_theory:
                row.append(me.theory_obtained if me.theory_obtained is not None else (me.special_value or ""))
            if subject.marking_structure.has_internal:
                row.append(me.internal_obtained if me.internal_obtained is not None else "")
                
            row.extend([
                sr_subj.theory_grade_point if sr_subj.theory_grade_point is not None else "",
                sr_subj.internal_grade_point if sr_subj.internal_grade_point is not None else "",
                sr_subj.grade or ""
            ])
            
        row.extend([
            sr.total_credit_hours if sr.total_credit_hours is not None else "",
            sr.overall_gpa if sr.overall_gpa is not None else "",
            sr.final_grade or "",
            sr.class_rank if sr.class_rank is not None else "",
            "Pass" if sr.is_pass else "Fail"
        ])
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{exam.name}_{cls.full_name}_Ledger.xlsx"'
    wb.save(response)
    return response


@login_required
def grade_ledger_select(request):
    """Grade Ledger selection page and view."""
    import json
    school = request.user.school
    from apps.exams.models import Exam, ExamClass
    from apps.classes.models import Class
    from apps.subjects.models import Subject
    from apps.marks.models import MarkEntry
    from apps.results.models import StudentResult

    active_session = school.get_active_session() if school else None

    # Fetch all exams
    exams = Exam.objects.filter(school=school)
    if active_session:
        exams = exams.filter(session=active_session)
    exams = exams.select_related('session').order_by('-created_at')

    # Build mapping of exam ID to its associated classes
    exam_classes_map = {}
    exam_classes = ExamClass.objects.filter(exam__school=school)
    if active_session:
        exam_classes = exam_classes.filter(exam__session=active_session)
    exam_classes = exam_classes.select_related('class_obj')
    for ec in exam_classes:
        if ec.is_locked:
            exam_classes_map.setdefault(ec.exam_id, []).append({
                'id': ec.class_obj.id,
                'name': ec.class_obj.full_name
            })

    exam_classes_json = json.dumps(exam_classes_map)

    # Check for selected parameters
    exam_id = request.GET.get('exam')
    class_id = request.GET.get('class_obj')

    context = {
        'exams': exams,
        'exam_classes_json': exam_classes_json,
        'selected_exam_id': int(exam_id) if exam_id and exam_id.isdigit() else None,
        'selected_class_id': int(class_id) if class_id and class_id.isdigit() else None,
        'school': school,
    }

    # If both parameters are supplied, fetch the grade ledger data
    if exam_id and class_id:
        try:
            exam = get_object_or_404(Exam, pk=exam_id, school=school)
            cls = get_object_or_404(Class, pk=class_id, school=school)

            from apps.exams.models import ExamClass
            exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=cls)
            if not exam_class.is_locked:
                messages.warning(request, f"Results for {cls.name} have not been processed and locked yet.")
                return redirect('grade_ledger_select')

            subjects = Subject.objects.filter(
                class_obj=cls, school=school
            ).select_related('marking_structure').order_by('order')

            student_results_qs = StudentResult.objects.filter(
                exam=exam, student__class_obj=cls, school=school
            ).select_related('student')
            
            def student_roll_sort_key(sr):
                try:
                    return int(sr.student.roll_number)
                except Exception:
                    return sr.student.roll_number
                    
            student_results = list(student_results_qs)
            student_results.sort(key=student_roll_sort_key)

            mark_map = {}
            for me in MarkEntry.objects.filter(
                exam=exam, school=school, student__class_obj=cls
            ).select_related('subject_result'):
                mark_map[(me.student_id, me.subject_id)] = me

            for sr in student_results:
                sr.subject_scores = []
                total_present = 0
                total_days = 0
                has_attendance = False

                for subject in subjects:
                    me = mark_map.get((sr.student_id, subject.id))
                    score_info = {
                        'subject': subject,
                        'has_theory': subject.marking_structure.has_theory if hasattr(subject, 'marking_structure') else True,
                        'theory_full_marks': subject.marking_structure.theory_full_marks if hasattr(subject, 'marking_structure') else 100,
                        'theory_obtained': '—',
                        'theory_grade': '—',
                        'has_internal': subject.marking_structure.has_internal if hasattr(subject, 'marking_structure') else False,
                        'internal_full_marks': subject.marking_structure.internal_full_marks if hasattr(subject, 'marking_structure') and subject.marking_structure.has_internal else 0,
                        'internal_obtained': '—',
                        'internal_grade': '—',
                    }
                    if me:
                        if me.present_days is not None:
                            total_present += me.present_days
                            has_attendance = True
                        if me.total_days:
                            total_days += me.total_days

                        if me.special_value:
                            val = me.get_special_value_display()
                            score_info['theory_obtained'] = val
                            score_info['theory_grade'] = val
                            score_info['theory_grade_point'] = val
                            if score_info['has_internal']:
                                score_info['internal_obtained'] = val
                                score_info['internal_grade'] = val
                                score_info['internal_grade_point'] = val
                            score_info['grade_point'] = val
                            score_info['grade'] = val
                        else:
                            if me.theory_obtained is not None:
                                score_info['theory_obtained'] = format_mark(me.theory_obtained)
                            if hasattr(me, 'subject_result') and me.subject_result:
                                score_info['theory_grade'] = me.subject_result.theory_grade or '—'
                                score_info['theory_grade_point'] = me.subject_result.theory_grade_point or 0.0
                            else:
                                score_info['theory_grade'] = '—'
                                score_info['theory_grade_point'] = '—'

                            if score_info['has_internal']:
                                if me.internal_obtained is not None:
                                    score_info['internal_obtained'] = format_mark(me.internal_obtained)
                                if hasattr(me, 'subject_result') and me.subject_result:
                                    score_info['internal_grade'] = me.subject_result.internal_grade or '—'
                                    score_info['internal_grade_point'] = me.subject_result.internal_grade_point or 0.0
                                else:
                                    score_info['internal_grade'] = '—'
                                    score_info['internal_grade_point'] = '—'

                            if hasattr(me, 'subject_result') and me.subject_result:
                                score_info['grade_point'] = me.subject_result.grade_point or 0.0
                                score_info['grade'] = me.subject_result.grade or '—'
                            else:
                                score_info['grade_point'] = '—'
                                score_info['grade'] = '—'
                    sr.subject_scores.append(score_info)

                if has_attendance and total_days > 0:
                    sr.attendance_pct = round((total_present / total_days) * 100, 1)
                else:
                    sr.attendance_pct = None

            context.update({
                'exam': exam,
                'class_obj': cls,
                'subjects': subjects,
                'student_results': student_results,
            })
        except Exception as e:
            messages.error(request, f"Error retrieving grade ledger: {str(e)}")

    return render(request, 'results/ledger_select.html', context)


@login_required
def marksheet_select(request):
    """Marksheet selection page and view."""
    import json
    school = request.user.school
    from apps.exams.models import Exam, ExamClass
    from apps.students.models import Student
    from apps.results.models import StudentResult
    from apps.marks.models import MarkEntry
    from apps.classes.models import Class

    active_session = school.get_active_session() if school else None

    # Fetch all exams
    exams = Exam.objects.filter(school=school)
    if active_session:
        exams = exams.filter(session=active_session)
    exams = exams.select_related('session').order_by('-created_at')

    # Build mapping of exam ID to its associated classes
    exam_classes_map = {}
    exam_classes = ExamClass.objects.filter(exam__school=school)
    if active_session:
        exam_classes = exam_classes.filter(exam__session=active_session)
    exam_classes = exam_classes.select_related('class_obj').order_by('class_obj__numeric_level', 'class_obj__name', 'class_obj__section').order_by('class_obj__numeric_level', 'class_obj__name', 'class_obj__section')
    for ec in exam_classes:
        if ec.is_locked:
            exam_classes_map.setdefault(ec.exam_id, []).append({
                'id': ec.class_obj.id,
                'name': ec.class_obj.full_name
            })

    # Build mapping of class ID to its students
    students_qs = Student.objects.filter(school=school, is_active=True)
    if active_session:
        students_qs = students_qs.filter(class_obj__session=active_session)
    students_qs = students_qs.order_by('roll_number')
    class_students_map = {}
    for s in students_qs:
        class_students_map.setdefault(s.class_obj_id, []).append({
            'id': s.id,
            'name': f"{s.name} (Roll: {s.roll_number})"
        })

    exam_classes_json = json.dumps(exam_classes_map)
    class_students_json = json.dumps(class_students_map)

    # Check for selected parameters
    exam_id = request.GET.get('exam')
    class_id = request.GET.get('class_obj')
    student_id = request.GET.get('student')
    
    if student_id == 'all':
        selected_student_id = 'all'
    else:
        selected_student_id = int(student_id) if student_id and student_id.isdigit() else None

    template_version = request.GET.get('template_version', 'default')

    context = {
        'exams': exams,
        'exam_classes_json': exam_classes_json,
        'class_students_json': class_students_json,
        'selected_exam_id': int(exam_id) if exam_id and exam_id.isdigit() else None,
        'selected_class_id': int(class_id) if class_id and class_id.isdigit() else None,
        'selected_student_id': selected_student_id,
        'template_version': template_version,
        'school': school,
    }

    # If all parameters are supplied, fetch the student marksheet data
    if exam_id and class_id and student_id:
        try:
            exam = get_object_or_404(Exam, pk=exam_id, school=school)
            cls = get_object_or_404(Class, pk=class_id, school=school)

            from apps.exams.models import ExamClass
            exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=cls)
            if not exam_class.is_locked:
                messages.warning(request, f"Results for {cls.name} have not been processed and locked yet.")
                return redirect('marksheet_select')

            if student_id == 'all':
                students_to_process = list(cls.students.filter(is_active=True).order_by('roll_number'))
                prev_student_id = None
                next_student_id = None
            else:
                student = get_object_or_404(Student, pk=student_id, school=school)
                students_to_process = [student]
                
                # Navigation - prev/next student relative to this class
                all_students = list(
                    student.class_obj.students.filter(is_active=True).order_by('roll_number')
                    .values_list('id', flat=True)
                )
                current_idx = all_students.index(student.id) if student.id in all_students else 0
                prev_student_id = all_students[current_idx - 1] if current_idx > 0 else None
                next_student_id = all_students[current_idx + 1] if current_idx < len(all_students) - 1 else None

            grade_remarks = {
                'A+': 'Outstanding',
                'A': 'Excellent',
                'B+': 'Very Good',
                'B': 'Good',
                'C+': 'Satisfactory',
                'C': 'Acceptable',
                'D+': 'Basic',
                'D': 'Basic',
                'E': 'Insufficient',
                'NG': 'Not Graded',
            }

            students_data = []

            for st in students_to_process:
                try:
                    result = StudentResult.objects.get(exam=exam, student=st, school=school)
                except StudentResult.DoesNotExist:
                    result = None

                mark_entries = MarkEntry.objects.filter(
                    exam=exam, student=st, school=school
                ).select_related(
                    'subject', 'subject__marking_structure', 'subject_result'
                ).order_by('subject__order')

                for me in mark_entries:
                    sr = getattr(me, 'subject_result', None)
                    if sr:
                        sr.theory_remark = grade_remarks.get(sr.theory_grade, '—') if sr.theory_grade else '—'
                        sr.internal_remark = grade_remarks.get(sr.internal_grade, '—') if sr.internal_grade else '—'
                        sr.overall_remark = grade_remarks.get(sr.grade, '—') if sr.grade else '—'

                present_days = '—'
                total_days = '—'
                for me in mark_entries:
                    if me.present_days is not None:
                        present_days = me.present_days
                    if me.total_days:
                        total_days = me.total_days
                    if present_days != '—' and total_days != '—':
                        break

                credit_mark_entries = [me for me in mark_entries if me.subject.subject_type != 'NON_CREDIT']
                non_credit_mark_entries = [me for me in mark_entries if me.subject.subject_type == 'NON_CREDIT']

                students_data.append({
                    'student': st,
                    'result': result,
                    'credit_mark_entries': credit_mark_entries,
                    'non_credit_mark_entries': non_credit_mark_entries,
                    'present_days': present_days,
                    'total_days': total_days,
                })

            context.update({
                'exam': exam,
                'class_obj': cls,
                'students_data': students_data,
                'prev_student_id': prev_student_id,
                'next_student_id': next_student_id,
            })
        except Exception:
            logger.exception('Error retrieving marksheet for exam=%s class=%s student=%s', exam_id, class_id, student_id)
            messages.error(request, 'An unexpected error occurred while generating the marksheet.')

    return render(request, 'results/marksheet_select.html', context)


@ratelimit(key='ip', rate='10/m', method='POST', block=True)
def public_result_search(request):
    """Public facing page for parents to search for a student's results."""
    from apps.students.models import Student
    from apps.results.models import StudentResult
    import datetime

    if request.method == 'POST':
        reg_no = request.POST.get('registration_number', '').strip()
        dob_str = request.POST.get('date_of_birth', '').strip()

        if not reg_no or not dob_str:
            messages.error(request, "Please provide both Registration Number and Date of Birth.")
            return redirect('public_result_search')

        # Find matching student by registration number
        students = Student.objects.filter(registration_number=reg_no)
        if not students.exists():
            messages.error(request, "No student found with the provided Registration Number.")
            return redirect('public_result_search')

        matched_student = None
        for s in students:
            if not s.date_of_birth:
                continue
            ad_dob_str = s.date_of_birth.strftime('%Y-%m-%d')
            bs_dob_str = s.dob_bs
            if dob_str == ad_dob_str or dob_str == bs_dob_str:
                matched_student = s
                break

        if not matched_student:
            messages.error(request, "Date of Birth does not match our records for this Registration Number.")
            return redirect('public_result_search')

        student = matched_student

        # SECURITY: Store auth in session with strict expiry (15 mins)
        request.session['auth_student_id'] = student.id
        request.session.set_expiry(900)

        
        # Get published results
        results = StudentResult.objects.filter(
            student=student, 
            exam__status='PUBLISHED' # Only published exams
        ).select_related('exam', 'exam__session').order_by('-exam__created_at')

        return render(request, 'results/public/search.html', {
            'student': student,
            'results': results,
            'searched': True
        })

    # GET request - check if session has a student auth to skip login
    if request.GET.get('logout') == '1':
        if 'auth_student_id' in request.session:
            del request.session['auth_student_id']
        messages.success(request, "Logged out successfully.")
        return redirect('public_result_search')

    student_id = request.session.get('auth_student_id')
    if student_id:
        student = Student.objects.filter(id=student_id).first()
        if student:
            results = StudentResult.objects.filter(
                student=student, 
                exam__status='PUBLISHED'
            ).select_related('exam', 'exam__session').order_by('-exam__created_at')
            
            return render(request, 'results/public/search.html', {
                'student': student,
                'results': results,
                'searched': True
            })

    return render(request, 'results/public/search.html')


def public_report_card(request, exam_id, student_id):
    """View a specific report card if authenticated via session."""
    from apps.exams.models import Exam
    from apps.students.models import Student
    from apps.results.models import StudentResult
    from apps.marks.models import MarkEntry

    # Verify session authentication
    auth_student_id = request.session.get('auth_student_id')
    if not auth_student_id or str(auth_student_id) != str(student_id):
        messages.error(request, "Unauthorized. Please enter the student's credentials again.")
        return redirect('public_result_search')

    student = get_object_or_404(Student, pk=student_id)
    exam = get_object_or_404(Exam, pk=exam_id, status='PUBLISHED')

    try:
        result = StudentResult.objects.get(exam=exam, student=student)
    except StudentResult.DoesNotExist:
        messages.error(request, "Result not found or not published.")
        return redirect('public_result_search')

    mark_entries = MarkEntry.objects.filter(
        exam=exam, student=student
    ).select_related(
        'subject', 'subject__marking_structure', 'subject_result'
    ).order_by('subject__order')

    grade_remarks = {
        'A+': 'Outstanding', 'A': 'Excellent', 'B+': 'Very Good', 'B': 'Good',
        'C+': 'Satisfactory', 'C': 'Acceptable', 'D+': 'Basic', 'D': 'Basic',
        'E': 'Insufficient', 'NG': 'Not Graded',
    }
    for me in mark_entries:
        sr = getattr(me, 'subject_result', None)
        if sr:
            sr.theory_remark = grade_remarks.get(sr.theory_grade, '—') if sr.theory_grade else '—'
            sr.internal_remark = grade_remarks.get(sr.internal_grade, '—') if sr.internal_grade else '—'
            sr.overall_remark = grade_remarks.get(sr.grade, '—') if sr.grade else '—'

    present_days = '—'
    total_days = '—'
    for me in mark_entries:
        if me.present_days is not None:
            present_days = me.present_days
        if me.total_days:
            total_days = me.total_days
        if present_days != '—' and total_days != '—':
            break

    credit_mark_entries = [me for me in mark_entries if me.subject.subject_type != 'NON_CREDIT']
    non_credit_mark_entries = [me for me in mark_entries if me.subject.subject_type == 'NON_CREDIT']

    return render(request, 'results/public/report_card.html', {
        'exam': exam,
        'student': student,
        'result': result,
        'credit_mark_entries': credit_mark_entries,
        'non_credit_mark_entries': non_credit_mark_entries,
        'present_days': present_days,
        'total_days': total_days,
    })
