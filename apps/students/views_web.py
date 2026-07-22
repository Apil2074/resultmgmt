"""
Students App — Web views with Excel import/export
"""
import io
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from .models import Student
from core.security import safe_redirect_url, validate_image_upload

logger = logging.getLogger(__name__)



@login_required
def student_list(request):
    school = request.user.school
    active_session = school.get_active_session() if school else None
    students = Student.objects.filter(school=school, is_active=True)
    if active_session:
        students = students.filter(class_obj__session=active_session)
    students = students.select_related(
        'class_obj', 'class_obj__session'
    )

    # Search and filter
    q = request.GET.get('q', '')
    class_id = request.GET.get('class_id', '')
    if q:
        students = students.filter(
            name__icontains=q
        ) | students.filter(roll_number__icontains=q)
    if class_id:
        students = students.filter(class_obj_id=class_id)

    # Convert to list and apply natural sorting (class, then roll number naturally)
    students_list = list(students)
    students_list.sort(key=lambda s: (
        s.class_obj.numeric_level,
        s.class_obj.name,
        s.class_obj.section,
        len(s.roll_number),
        s.roll_number
    ))

    from apps.classes.models import Class
    classes_qs = Class.objects.filter(school=school)
    if active_session:
        classes_qs = classes_qs.filter(session=active_session)
    classes = list(classes_qs)

    paginator = Paginator(students_list, 50)  # 50 students per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'students/list.html', {
        'page_obj': page_obj,
        'classes': classes,
        'q': q,
        'class_id': class_id,
    })


@login_required
def student_create(request):
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        if request.user.is_teacher:
            return redirect('teacher_dashboard')
        return redirect('student_list')
    school = request.user.school
    active_session = school.get_active_session() if school else None
    if not active_session:
        messages.error(request, 'You must create an active academic session before adding students.')
        return redirect('student_list')
    from apps.classes.models import Class
    classes = Class.objects.filter(school=school)
    if active_session:
        classes = classes.filter(session=active_session)

    if request.method == 'POST':
        if school.is_demo and Student.objects.filter(school=school).count() >= 10:
            messages.error(request, 'Demo accounts are limited to adding 10 students.')
            return redirect('student_list')
            
        Student.objects.create(
            school=school,
            class_obj_id=request.POST.get('class_id'),
            roll_number=request.POST.get('roll_number'),
            name=request.POST.get('name'),
            gender=request.POST.get('gender'),
            registration_number=request.POST.get('registration_number', ''),
            symbol_number=request.POST.get('symbol_number', ''),
            date_of_birth=request.POST.get('date_of_birth') or None,
            parent_name=request.POST.get('parent_name', ''),
            contact_number=request.POST.get('contact_number', ''),
            address=request.POST.get('address', ''),
            photo=request.FILES.get('photo'),
        )
        messages.success(request, 'Student added successfully.')
        return redirect('student_list')

    return render(request, 'students/form.html', {'classes': classes})


@login_required
def student_detail(request, pk):
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school)
    from apps.results.models import StudentResult
    results = StudentResult.objects.filter(
        student=student
    ).select_related('exam').order_by('-exam__start_date')
    return render(request, 'students/detail.html', {
        'student': student,
        'results': results,
    })


@login_required
def student_edit(request, pk):
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        if request.user.is_teacher:
            return redirect('teacher_dashboard')
        return redirect('student_list')
    school = request.user.school
    active_session = school.get_active_session() if school else None
    student = get_object_or_404(Student, pk=pk, school=school)
    from apps.classes.models import Class
    classes = Class.objects.filter(school=school)
    if active_session:
        classes = classes.filter(session=active_session)

    if request.method == 'POST':
        student.name = request.POST.get('name', student.name)
        student.roll_number = request.POST.get('roll_number', student.roll_number)
        student.registration_number = request.POST.get('registration_number', '')
        student.symbol_number = request.POST.get('symbol_number', '')
        student.gender = request.POST.get('gender', student.gender)
        student.parent_name = request.POST.get('parent_name', '')
        student.contact_number = request.POST.get('contact_number', '')
        student.address = request.POST.get('address', '')
        dob = request.POST.get('date_of_birth')
        student.date_of_birth = dob if dob else None
        if 'photo' in request.FILES:
            from django.core.exceptions import ValidationError
            try:
                validate_image_upload(request.FILES['photo'])
                student.photo = request.FILES['photo']
            except ValidationError as e:
                messages.error(request, str(e.message))
                return render(request, 'students/form.html', {'student': student, 'classes': classes})
        elif request.POST.get('remove_photo'):
            if student.photo:
                student.photo.delete(save=False)
                student.photo = None

        student.save()
        messages.success(request, 'Student updated.')

        next_url = safe_redirect_url(
            request.POST.get('next') or request.GET.get('next'), request
        )
        if next_url:
            return redirect(next_url)
        return redirect('student_detail', pk=student.pk)

    return render(request, 'students/form.html', {
        'student': student,
        'classes': classes,
    })


@login_required
def student_delete(request, pk):
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        if request.user.is_teacher:
            return redirect('teacher_dashboard')
        return redirect('student_list')
    school = request.user.school
    student = get_object_or_404(Student, pk=pk, school=school)
    if request.method == 'POST':
        student_name = student.name
        student.delete()
        messages.success(request, f'Student {student_name} permanently deleted.')

        # SECURITY: Validate next= to prevent open redirect attacks
        next_url = safe_redirect_url(
            request.POST.get('next') or request.GET.get('next'), request
        )
        return redirect(next_url or 'student_list')
    return render(request, 'students/confirm_delete.html', {'student': student})


def parse_class_and_section(class_str):
    class_str = class_str.strip()
    if not class_str:
        return None, None
    import re
    # Match Class 10 A, Class 10 - A, Class 10A, Grade 9 B, etc.
    match = re.search(r'^(.*?)\s*[- ]\s*([a-zA-Z])$', class_str)
    if match:
        return match.group(1).strip(), match.group(2).strip().upper()
    return class_str, ''


@login_required
def student_import(request):
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        if request.user.is_teacher:
            return redirect('teacher_dashboard')
        return redirect('student_list')
    """Import students from Excel."""
    school = request.user.school
    active_session = school.get_active_session() if school else None
    if not active_session:
        messages.error(request, 'You must create an active academic session before importing students.')
        return redirect('student_list')
    from apps.classes.models import Class
    classes = Class.objects.filter(school=school)
    if active_session:
        classes = classes.filter(session=active_session)
        
    if not classes.exists():
        messages.error(request, 'Please create a class first before importing students.')
        return redirect('class_list')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl
        import datetime
        from django.utils import timezone
        
        excel_file = request.FILES['excel_file']
        class_id = request.POST.get('class_id')
        dropdown_cls = Class.objects.filter(pk=class_id, school=school).first() if class_id else None

        active_session = school.get_active_session()
        if not active_session:
            from apps.schools.models import AcademicSession
            active_session, _ = AcademicSession.objects.get_or_create(
                school=school,
                is_active=True,
                defaults={'name': str(timezone.now().year)}
            )

        try:
            wb = openpyxl.load_workbook(excel_file)
            created = 0
            errors = []

            # Track next roll number per class (class pk -> next int)
            class_roll_counters = {}

            def get_next_roll(cls_obj):
                """Return the next sequential roll number for this class."""
                if cls_obj.pk not in class_roll_counters:
                    existing = Student.objects.filter(
                        school=school, class_obj=cls_obj
                    ).values_list('roll_number', flat=True)
                    max_roll = 0
                    for r in existing:
                        try:
                            max_roll = max(max_roll, int(r))
                        except (ValueError, TypeError):
                            pass
                    class_roll_counters[cls_obj.pk] = max_roll
                class_roll_counters[cls_obj.pk] += 1
                return str(class_roll_counters[cls_obj.pk])

            import datetime
            from apps.classes.models import Class
            for ws in wb.worksheets:
                class_name = ws.title
                if class_name.lower() == 'students' and not Class.objects.filter(school=school, name__iexact=class_name).exists():
                    # It might be the default sheet, fallback to dropdown_cls if provided
                    cls = dropdown_cls
                else:
                    found_cls = Class.objects.filter(school=school, name__iexact=class_name).first()
                    if not found_cls:
                        for c in Class.objects.filter(school=school):
                            if c.full_name.lower() == class_name.lower():
                                found_cls = c
                                break
                    
                    cls = found_cls
                
                # If no class could be matched and no dropdown class, skip
                if not cls:
                    errors.append(f"Sheet '{class_name}': No class found matching '{class_name}' and no default class selected.")
                    continue

                for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                    if not row[0]:
                        continue
                    def safe_str(val):
                        if val is None: return ""
                        if isinstance(val, float) and val.is_integer():
                            return str(int(val))
                        return str(val).strip()
                    
                    sn_val = safe_str(row[0]).upper()
                    if (sn_val == school.name.upper() or 'ACADEMIC YEAR:' in sn_val or 
                        sn_val.startswith('CLASS:') or sn_val in ('SN*', 'SN', 'ROLL NO*', 'ROLL NO')):
                        continue

                    try:
                        symbol = safe_str(row[1]) if len(row) > 1 else ''
                        reg_num = safe_str(row[2]) if len(row) > 2 else ''
                        name = safe_str(row[3]) if len(row) > 3 else ''
                        if not name:
                            raise ValueError("Student name is required.")
                        
                        gender_raw = safe_str(row[4]) if len(row) > 4 else ''
                        gender = None
                        if gender_raw:
                            gen_cap = gender_raw.upper()
                            if gen_cap in ('M', 'MALE'): gender = 'M'
                            elif gen_cap in ('F', 'FEMALE'): gender = 'F'
                            elif gen_cap in ('O', 'OTHER'): gender = 'O'
                        
                        dob_val = row[5] if len(row) > 5 else None
                        parent_name = safe_str(row[6]) if len(row) > 6 else ''
                        contact_number = safe_str(row[7]) if len(row) > 7 else ''
                        address = safe_str(row[8]) if len(row) > 8 else ''
                        
                        dob_ad = None
                        dob_bs_str = ""
                        
                        if dob_val is not None:
                            import nepali_datetime
                            if isinstance(dob_val, (datetime.date, datetime.datetime)):
                                y, m, d = dob_val.year, dob_val.month, dob_val.day
                                dob_bs_str = f"{y:04d}-{m:02d}-{d:02d}"
                                try:
                                    np_date = nepali_datetime.date(y, m, d)
                                    dob_ad = np_date.to_datetime_date()
                                except Exception:
                                    pass
                            else:
                                dob_str = str(dob_val).strip()
                                if dob_str:
                                    dob_bs_str = dob_str
                                    try:
                                        parts = [int(p) for p in dob_str.split('-')]
                                        if len(parts) == 3:
                                            y, m, d = parts
                                            np_date = nepali_datetime.date(y, m, d)
                                            dob_ad = np_date.to_datetime_date()
                                    except Exception:
                                        pass

                        student = None
                        if symbol:
                            student = Student.objects.filter(school=school, symbol_number=symbol).first()
                        if not student and reg_num:
                            student = Student.objects.filter(school=school, registration_number=reg_num).first()
                            
                        sn_val_raw = safe_str(row[0])
                        if not student and sn_val_raw:
                            student = Student.objects.filter(school=school, class_obj=cls, roll_number=sn_val_raw, name__iexact=name).first()
                        
                        if student:
                            if name: student.name = name
                            if symbol: student.symbol_number = symbol
                            if reg_num: student.registration_number = reg_num
                            if gender: student.gender = gender
                            if cls and student.class_obj != cls:
                                student.class_obj = cls
                                student.roll_number = get_next_roll(cls)
                            elif cls:
                                student.class_obj = cls
                            if dob_ad: student.date_of_birth = dob_ad
                            if dob_bs_str: student.date_of_birth_bs = dob_bs_str
                            if parent_name: student.parent_name = parent_name
                            if contact_number: student.contact_number = contact_number
                            if address: student.address = address
                            student.save()
                        else:
                            if school.is_demo and Student.objects.filter(school=school).count() >= 10:
                                raise ValueError("Demo accounts are limited to 10 students. Import stopped.")
                            Student.objects.create(
                                school=school,
                                name=name,
                                symbol_number=symbol,
                                registration_number=reg_num,
                                gender=gender,
                                class_obj=cls,
                                roll_number=get_next_roll(cls),
                                date_of_birth=dob_ad,
                                date_of_birth_bs=dob_bs_str,
                                parent_name=parent_name,
                                contact_number=contact_number,
                                address=address,
                                is_active=True
                            )
                            created += 1

                    except Exception as e:
                        errors.append(f"Sheet '{ws.title}' Row {row_num}: {str(e)}")

            if errors:
                messages.warning(request, f"Imported {created} students. However, there were some errors: {', '.join(errors[:5])}{'...' if len(errors) > 5 else ''}")
            else:
                messages.success(request, f"Successfully processed students. {created} new created, others updated.")
        except Exception as e:
            messages.error(request, f'Failed to read Excel file: {str(e)}')

        return redirect('student_list')

    return render(request, 'students/import.html', {'classes': classes})


@login_required
def student_export_excel(request):
    """Export student list to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    school = request.user.school
    active_session = school.get_active_session() if school else None
    class_id = request.GET.get('class_id')

    students = Student.objects.filter(school=school, is_active=True)
    if active_session:
        students = students.filter(class_obj__session=active_session)
    if class_id:
        students = students.filter(class_obj_id=class_id)
    students = students.select_related('class_obj')

    wb = openpyxl.Workbook()
    default_ws = wb.active
    
    headers = ['Roll No*', 'Symbol No', 'REG NO', 'Name of Students*', 'Gender', 'Date of Birth BS.', 'Parent Name', 'Contact Number', 'Address']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')

    from collections import defaultdict
    class_students = defaultdict(list)
    for student in students:
        if student.class_obj:
            class_students[student.class_obj.name].append(student)
        else:
            class_students['Unknown'].append(student)

    if not class_students:
        default_ws.title = 'Students'
        for col, header in enumerate(headers, 1):
            cell = default_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
    else:
        for i, (cls_name, stu_list) in enumerate(class_students.items()):
            import re
            safe_title = re.sub(r'[\\*?:/\[\]]', '', cls_name)[:31]
            if i == 0:
                ws = default_ws
                ws.title = safe_title
            else:
                ws = wb.create_sheet(title=safe_title)
                
            # Beautiful Headers
            ws.merge_cells('B1:G3')
            details = []
            if active_session:
                details.append(f"ACADEMIC YEAR: {active_session.name}")
            details.append(f"CLASS: {safe_title}")
            header_text = f"{school.name.upper()}\n{' | '.join(details)}"
            cell_header = ws.cell(row=1, column=2, value=header_text)
            cell_header.font = Font(bold=True, size=14, color='1E293B')
            cell_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Remove gridlines for the header rows
            for r in range(1, 4):
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row_num, student in enumerate(stu_list, 5):
                ws.cell(row=row_num, column=1, value=student.roll_number or row_num - 4)
                ws.cell(row=row_num, column=2, value=student.symbol_number)
                ws.cell(row=row_num, column=3, value=student.registration_number)
                ws.cell(row=row_num, column=4, value=student.name)
                ws.cell(row=row_num, column=5, value=student.get_gender_display() or '')
                ws.cell(row=row_num, column=6, value=student.date_of_birth_bs or '')
                ws.cell(row=row_num, column=7, value=student.parent_name)
                ws.cell(row=row_num, column=8, value=student.contact_number)
                ws.cell(row=row_num, column=9, value=student.address)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    from openpyxl.utils import get_column_letter
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=4, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max_len + 6, 40)

    from apps.classes.models import Class
    target_class = Class.objects.filter(pk=class_id).first() if class_id else None
    filename = f"{target_class.name} Students.xlsx" if target_class else "Students.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def student_import_template(request):
    """Download blank Excel template for student import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    school = request.user.school
    active_session = school.get_active_session() if school else None
    
    from apps.classes.models import Class
    classes = Class.objects.filter(school=school, session=active_session) if active_session else []
    
    wb = openpyxl.Workbook()
    default_ws = wb.active
    
    headers = ['Roll No*', 'Symbol No', 'REG NO', 'Name of Students*', 'Gender', 'Date of Birth BS.', 'Parent Name', 'Contact Number', 'Address']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    
    def setup_sheet(ws, title):
        import re
        safe_title = re.sub(r'[\\*?:/\[\]]', '', title)[:31]
        ws.title = safe_title
        
        # Beautiful Headers
        ws.merge_cells('B1:G3')
        details = []
        if active_session:
            details.append(f"ACADEMIC YEAR: {active_session.name}")
        details.append(f"CLASS: {safe_title}")
        header_text = f"{school.name.upper()}\n{' | '.join(details)}"
        cell_header = ws.cell(row=1, column=2, value=header_text)
        cell_header.font = Font(bold=True, size=14, color='1E293B')
        cell_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Remove gridlines for the header rows
        for r in range(1, 4):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            
        for r in range(4, 34):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = thin_border
        
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 6, 40)

    if classes:
        for i, cls in enumerate(classes):
            if i == 0:
                ws = default_ws
            else:
                ws = wb.create_sheet()
            setup_sheet(ws, cls.name)
    else:
        setup_sheet(default_ws, 'Students')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    wb.save(response)
    return response

@login_required
def student_bulk_delete(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if student_ids:
            school = request.user.school
            students = Student.objects.filter(id__in=student_ids, school=school)
            count = students.count()
            students.delete()
        else:
            messages.warning(request, "No students selected for deletion.")
    return redirect(request.META.get('HTTP_REFERER', 'student_list'))

@login_required
def student_inline_edit(request, pk):
    from django.http import JsonResponse
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        return JsonResponse({'status': 'error', 'message': 'Access denied.'}, status=403)
    
    if request.method == 'POST':
        school = request.user.school
        student = get_object_or_404(Student, pk=pk, school=school)
        
        field = request.POST.get('name')
        value = request.POST.get('value')
        
        allowed_fields = [
            'name', 'roll_number', 'registration_number', 
            'symbol_number', 'contact_number', 'gender',
            'parent_name', 'address', 'date_of_birth_bs'
        ]
        
        if field in allowed_fields:
            if field == 'date_of_birth_bs':
                student.date_of_birth_bs = value
            else:
                setattr(student, field, value)
                
            try:
                student.save()
                return JsonResponse({'status': 'success', 'message': 'Updated successfully'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        
        return JsonResponse({'status': 'error', 'message': 'Invalid field'}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@login_required
def student_edit_modal(request, pk):
    from django.http import JsonResponse
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        return JsonResponse({'status': 'error', 'message': 'Access denied.'}, status=403)
        
    school = request.user.school
    active_session = school.get_active_session() if school else None
    student = get_object_or_404(Student, pk=pk, school=school)
    
    from apps.classes.models import Class
    classes = Class.objects.filter(school=school)
    if active_session:
        classes = classes.filter(session=active_session)
        
    if request.method == 'GET':
        return render(request, 'students/form_partial.html', {
            'student': student,
            'classes': classes,
            'is_modal': True
        })
        
    if request.method == 'POST':
        student.name = request.POST.get('name', student.name)
        student.roll_number = request.POST.get('roll_number', student.roll_number)
        student.class_obj_id = request.POST.get('class_id', student.class_obj_id)
        student.registration_number = request.POST.get('registration_number', '')
        student.symbol_number = request.POST.get('symbol_number', '')
        student.gender = request.POST.get('gender', student.gender)
        student.parent_name = request.POST.get('parent_name', '')
        student.contact_number = request.POST.get('contact_number', '')
        student.address = request.POST.get('address', '')
        
        dob = request.POST.get('date_of_birth')
        if dob:
            student.date_of_birth = dob
            student.date_of_birth_bs = dob
        else:
            student.date_of_birth = None
            student.date_of_birth_bs = None
            
        if 'photo' in request.FILES:
            from django.core.exceptions import ValidationError
            from core.security import validate_image_upload
            try:
                validate_image_upload(request.FILES['photo'])
                student.photo = request.FILES['photo']
            except ValidationError as e:
                return JsonResponse({'status': 'error', 'message': str(e.message)}, status=400)
        elif request.POST.get('remove_photo'):
            if student.photo:
                student.photo.delete(save=False)
                student.photo = None
                
        try:
            student.save()
            
            # Return updated data for dynamic UI update
            data = {
                'name': student.name,
                'roll_number': student.roll_number,
                'class_name': student.class_obj.full_name if student.class_obj else '',
                'registration_number': student.registration_number or '—',
                'symbol_number': student.symbol_number or '',
                'contact_number': student.contact_number or '—',
                'photo_url': student.photo_url,
                'dob_bs': student.dob_bs or ''
            }
            return JsonResponse({'status': 'success', 'message': 'Student updated successfully.', 'data': data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)
