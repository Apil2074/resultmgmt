"""
Marks App — Web views for mark entry and bulk import/export
"""
import io
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction, IntegrityError
import json
from .models import MarkEntry
from core.security import validate_excel_upload

logger = logging.getLogger(__name__)


@login_required
def mark_entry(request, exam_id, class_id):
    """Mark entry page — inline table for entering marks."""
    school = request.user.school
    from apps.exams.models import Exam
    from apps.classes.models import Class
    from apps.students.models import Student
    from apps.subjects.models import Subject

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    cls = get_object_or_404(Class, pk=class_id, school=school)
    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=cls)

    if exam_class.is_locked:
        if request.user.is_teacher:
            messages.warning(request, 'Mark entry is locked by Admin.')
        else:
            messages.warning(request, 'Results for this class are locked and published. Unlock to edit marks.')

    students = Student.objects.filter(class_obj=cls, school=school, is_active=True)
    subjects = Subject.objects.filter(class_obj=cls, school=school).select_related().order_by('order')

    is_class_teacher = False
    assigned_subject_ids = []
    
    if request.user.is_teacher:
        if hasattr(request.user, 'teacher_profile'):
            teacher_profile = request.user.teacher_profile
            is_class_teacher = (cls.class_teacher == teacher_profile)
            assigned_subject_ids = list(teacher_profile.subject_assignments.filter(subject__class_obj=cls).values_list('subject_id', flat=True))
            
            if not is_class_teacher and not assigned_subject_ids:
                messages.error(request, 'You do not have access to enter marks for this class.')
                return redirect('dashboard')
        else:
            messages.error(request, 'Teacher profile not found.')
            return redirect('dashboard')
    else:
        is_class_teacher = True

    # For subject-only teachers, restrict visible subjects to only assigned ones
    if request.user.is_teacher and not is_class_teacher:
        subjects = subjects.filter(id__in=assigned_subject_ids)

    # Load optional enrollments
    from apps.subjects.models import StudentSubjectEnrollment, Subject
    enrolled_pairs = set(
        StudentSubjectEnrollment.objects.filter(
            student__in=students,
            subject__in=subjects
        ).values_list('student_id', 'subject_id')
    )

    # Load existing mark entries
    existing_marks = {}
    for me in MarkEntry.objects.filter(exam=exam, school=school,
                                        student__in=students, subject__in=subjects):
        existing_marks[(me.student_id, me.subject_id)] = me

    # Enrich student objects for the template
    for student in students:
        student.subject_marks = []
        for subject in subjects:
            me = existing_marks.get((student.id, subject.id))
            
            is_unmapped_optional = (
                subject.subject_type == Subject.SubjectType.OPTIONAL and
                (student.id, subject.id) not in enrolled_pairs
            )
            
            student.subject_marks.append({
                'subject': subject,
                'me': me,
                'is_restricted': is_unmapped_optional,
            })
        
        # Load attendance from any existing mark entry of this student
        present_days = ''
        total_days = ''
        for me in existing_marks.values():
            if me.student_id == student.id:
                if me.present_days is not None:
                    present_days = me.present_days
                if me.total_days:
                    total_days = me.total_days
                break
        student.present_days_val = present_days
        student.total_days_val = total_days

    # Get all classes for this exam to show in the dropdown
    all_exam_classes = ExamClass.objects.filter(exam=exam).select_related('class_obj')
    available_classes = [ec.class_obj for ec in all_exam_classes]

    return render(request, 'marks/entry.html', {
        'exam': exam,
        'class_obj': cls,
        'exam_class': exam_class,
        'students': students,
        'subjects': subjects,
        'available_classes': available_classes,
        'is_class_teacher': is_class_teacher,
    })


@login_required
def save_mark(request):
    """AJAX endpoint to save a single mark entry."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    school = request.user.school
    from apps.exams.models import Exam
    from apps.students.models import Student
    from apps.subjects.models import Subject

    exam = get_object_or_404(Exam, pk=data.get('exam_id'), school=school)
    student = get_object_or_404(Student, pk=data.get('student_id'), school=school)
    subject = get_object_or_404(Subject, pk=data.get('subject_id'), school=school)

    if subject.class_obj_id != student.class_obj_id:
        return JsonResponse({'error': "Subject does not belong to the student's class"}, status=400)

    from apps.exams.models import ExamClass
    exam_class = get_object_or_404(ExamClass, exam=exam, class_obj=student.class_obj)
    if exam_class.is_locked:
        return JsonResponse({'error': 'Exam is locked for this class'}, status=403)

    if request.user.is_teacher:
        if not hasattr(request.user, 'teacher_profile'):
            return JsonResponse({'error': 'Unauthorized'}, status=403)
            
        teacher_profile = request.user.teacher_profile
        is_class_teacher = (subject.class_obj.class_teacher == teacher_profile)
        is_assigned_subject = teacher_profile.subject_assignments.filter(subject=subject).exists()
        
        if not is_class_teacher and not is_assigned_subject:
            return JsonResponse({'error': 'Unauthorized'}, status=403)

    if subject.subject_type == Subject.SubjectType.OPTIONAL:
        from apps.subjects.models import StudentSubjectEnrollment
        if not StudentSubjectEnrollment.objects.filter(student=student, subject=subject).exists():
            return JsonResponse({'error': 'Cannot enter marks for an unmapped optional subject.'}, status=400)

    special_value = data.get('special_value') or None
    from decimal import Decimal

    try:
        theory_val = data.get('theory_obtained')
        theory_obtained = Decimal(theory_val) if theory_val not in (None, '') else None
    except Exception:
        return JsonResponse({'error': 'Invalid theory marks value'}, status=400)

    try:
        internal_val = data.get('internal_obtained')
        internal_obtained = Decimal(internal_val) if internal_val not in (None, '') else None
    except Exception:
        return JsonResponse({'error': 'Invalid internal marks value'}, status=400)

    try:
        pres_days = data.get('present_days')
        present_days = int(pres_days) if pres_days not in (None, '') else None
    except Exception:
        return JsonResponse({'error': 'Invalid present days'}, status=400)

    try:
        tot_days = data.get('total_days')
        total_days = int(tot_days) if tot_days not in (None, '') else None
    except Exception:
        return JsonResponse({'error': 'Invalid total days'}, status=400)

    defaults = {
        'school': school,
        'special_value': special_value,
        'theory_obtained': theory_obtained,
        'internal_obtained': internal_obtained,
        'entered_by': request.user,
    }
    if not getattr(request.user, 'is_teacher', False):
        defaults['present_days'] = present_days
        defaults['total_days'] = total_days

    try:
        with transaction.atomic():
            me, created = MarkEntry.objects.select_for_update().get_or_create(
                exam=exam, student=student, subject=subject,
                defaults=defaults
            )
            if not created:
                for k, v in defaults.items():
                    setattr(me, k, v)
                me.save()
        return JsonResponse({
            'status': 'saved',
            'id': me.pk,
            'total_obtained': str(me.total_obtained) if me.total_obtained is not None else None,
        })
    except IntegrityError:
        return JsonResponse({'error': 'Mark entry conflict. Please refresh and try again.'}, status=409)
    except Exception:
        logger.exception('Unexpected error in save_mark for exam=%s student=%s subject=%s',
                         data.get('exam_id'), data.get('student_id'), data.get('subject_id'))
        return JsonResponse({'error': 'An unexpected error occurred. Please try again.'}, status=500)


@login_required
def save_marks_bulk(request):
    """AJAX endpoint to save multiple mark entries at once."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data_list = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if not isinstance(data_list, list):
        return JsonResponse({'error': 'Payload must be an array of marks'}, status=400)

    school = request.user.school
    from apps.exams.models import Exam
    from apps.students.models import Student
    from apps.subjects.models import Subject
    from decimal import Decimal
    from apps.marks.models import MarkEntry

    if not data_list:
        return JsonResponse({'status': 'saved', 'count': 0})
        
    exam_id = data_list[0].get('exam_id')
    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    
    student_ids = [d.get('student_id') for d in data_list]
    students = {s.id: s for s in Student.objects.filter(id__in=student_ids, school=school)}
    
    if students:
        from apps.exams.models import ExamClass
        # SECURITY: Check lock for EVERY student's class, not just the first.
        # A single locked class check can be bypassed by putting an unlocked student first.
        checked_classes = set()
        for data_item in data_list:
            sid = data_item.get('student_id')
            s = students.get(sid)
            if not s:
                continue
            class_id = s.class_obj_id
            if class_id in checked_classes:
                continue
            checked_classes.add(class_id)
            exam_class = ExamClass.objects.filter(exam=exam, class_obj=s.class_obj).first()
            if exam_class and exam_class.is_locked:
                return JsonResponse(
                    {'error': f'Exam is locked for class: {s.class_obj.name}. Unlock before entering marks.'},
                    status=403
                )

    is_teacher = getattr(request.user, 'is_teacher', False)

    student_ids = [d.get('student_id') for d in data_list]
    subject_ids = [d.get('subject_id') for d in data_list]
    
    students = {s.id: s for s in Student.objects.filter(id__in=student_ids, school=school)}
    subjects = {s.id: s for s in Subject.objects.filter(id__in=subject_ids, school=school)}
    
    if request.user.is_teacher:
        if not hasattr(request.user, 'teacher_profile'):
            return JsonResponse({'error': 'Unauthorized'}, status=403)
        assigned_subjects = set(Subject.objects.filter(id__in=subject_ids, class_obj__class_teacher=request.user.teacher_profile).values_list('id', flat=True))

    from apps.subjects.models import StudentSubjectEnrollment
    optional_enrollments = set(StudentSubjectEnrollment.objects.filter(
        student_id__in=student_ids, subject_id__in=subject_ids
    ).values_list('student_id', 'subject_id'))

    updates = []
    creates = []
    
    existing_marks = {(m.student_id, m.subject_id): m for m in MarkEntry.objects.filter(exam=exam, student_id__in=student_ids, subject_id__in=subject_ids)}

    for data in data_list:
        student_id = data.get('student_id')
        subject_id = data.get('subject_id')
        student = students.get(student_id)
        subject = subjects.get(subject_id)
        
        if not student or not subject:
            return JsonResponse({'error': f'Invalid student or subject ID: {student_id}, {subject_id}'}, status=400)

        if subject.class_obj_id != student.class_obj_id:
            return JsonResponse({'error': f'Subject {subject.name} does not belong to the class of {student.name}.'}, status=400)

        if is_teacher and subject.id not in assigned_subjects:
            return JsonResponse({'error': f'You are not assigned to subject: {subject.name}.'}, status=403)

        if subject.subject_type == Subject.SubjectType.OPTIONAL:
            if (student.id, subject.id) not in optional_enrollments:
                return JsonResponse({'error': f'Cannot enter marks for unmapped optional subject: {subject.name} for {student.name}.'}, status=400)

        special_value = data.get('special_value') or None
        
        try:
            theory_val = data.get('theory_obtained')
            theory_obtained = Decimal(str(theory_val)) if theory_val not in (None, '') else None
        except Exception:
            return JsonResponse({'error': f'Invalid theory marks value for {student.name}'}, status=400)

        try:
            internal_val = data.get('internal_obtained')
            internal_obtained = Decimal(str(internal_val)) if internal_val not in (None, '') else None
        except Exception:
            return JsonResponse({'error': f'Invalid internal marks value for {student.name}'}, status=400)

        try:
            pres_days = data.get('present_days')
            present_days = int(pres_days) if pres_days not in (None, '') else None
        except Exception:
            return JsonResponse({'error': f'Invalid present days for {student.name}'}, status=400)

        try:
            tot_days = data.get('total_days')
            total_days = int(tot_days) if tot_days not in (None, '') else None
        except Exception:
            return JsonResponse({'error': f'Invalid total days for {student.name}'}, status=400)
            
        me = existing_marks.get((student.id, subject.id))
        if me:
            me.special_value = special_value
            me.theory_obtained = theory_obtained
            me.internal_obtained = internal_obtained
            me.entered_by = request.user
            if not is_teacher:
                me.present_days = present_days
                me.total_days = total_days
            updates.append(me)
        else:
            me = MarkEntry(
                school=school,
                exam=exam,
                session=exam.session,
                student=student,
                subject=subject,
                special_value=special_value,
                theory_obtained=theory_obtained,
                internal_obtained=internal_obtained,
                entered_by=request.user,
            )
            if not is_teacher:
                me.present_days = present_days
                me.total_days = total_days
            creates.append(me)

    try:
        if creates:
            for c in creates:
                c.clean()
            MarkEntry.objects.bulk_create(creates)
        if updates:
            for u in updates:
                u.clean()
            MarkEntry.objects.bulk_update(updates, ['special_value', 'theory_obtained', 'internal_obtained', 'entered_by', 'present_days', 'total_days'])

        return JsonResponse({'status': 'saved', 'count': len(creates) + len(updates)})
    except Exception as e:
        logger.exception('Unexpected error in save_marks_bulk for exam=%s', exam_id)
        return JsonResponse({'error': f'An unexpected error occurred while saving marks: {str(e)}'}, status=500)


@login_required
def bulk_mark_import(request, exam_id, class_id):
    """Bulk Excel mark import."""
    school = request.user.school
    from apps.exams.models import Exam
    from apps.classes.models import Class
    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    cls = get_object_or_404(Class, pk=class_id, school=school)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        # SECURITY: Validate file type and size before processing
        from django.core.exceptions import ValidationError
        try:
            validate_excel_upload(request.FILES['excel_file'])
        except ValidationError as ve:
            messages.error(request, str(ve.message))
            return redirect('mark_entry', exam_id=exam_id, class_id=class_id)

        import openpyxl
        from apps.students.models import Student
        from apps.subjects.models import Subject, StudentSubjectEnrollment

        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            ws = wb.active
            # Read header row to get subject IDs
            subjects = list(Subject.objects.filter(class_obj=cls, school=school).order_by('order'))
            students_qs = Student.objects.filter(class_obj=cls, school=school)
            students = {s.roll_number: s for s in students_qs}
            
            optional_enrollments = set(StudentSubjectEnrollment.objects.filter(
                student__in=students_qs, subject__in=subjects
            ).values_list('student_id', 'subject_id'))

            # Validate headers to prevent users from reordering columns
            expected_headers = ['Roll No', 'Student Name']
            for subj in subjects:
                try:
                    ms = subj
                    expected_headers.append(f'{subj.name}\nTheory')
                    if ms.has_practical:
                        expected_headers.append(f'{subj.name}\nInternal')
                    else:
                        expected_headers.append(f'{subj.name}\n(No Internal)')
                except Exception:
                    expected_headers.append(subj.name)
                    expected_headers.append('')
            expected_headers.append('Present Days')
            
            actual_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            actual_headers_clean = [str(h).strip() if h is not None else '' for h in actual_headers][:len(expected_headers)]
            expected_headers_clean = [str(h).strip() for h in expected_headers]
            
            if actual_headers_clean != expected_headers_clean:
                messages.error(request, 'Column structure mismatch detected. Please download a fresh template and do not alter the headers or column order.')
                return redirect('mark_entry', exam_id=exam_id, class_id=class_id)

            # Check if Row 2 is the Full Marks / Total config row
            row_2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            has_config_row = False
            class_total_days = None
            
            if len(row_2) > 1 and row_2[1] and any(x in str(row_2[1]).lower() for x in ('full marks', 'total')):
                has_config_row = True
                
            if has_config_row:
                # 1. Parse and update subject full marks
                col_idx = 2
                for subject in subjects:
                    try:
                        ms = subject
                        changed = False
                        
                        theory_fm = row_2[col_idx] if col_idx < len(row_2) else None
                        if theory_fm is not None:
                            try:
                                val = int(float(theory_fm))
                                if ms.theory_full_marks != val:
                                    ms.theory_full_marks = val
                                    changed = True
                            except Exception:
                                pass
                                
                        if ms.has_practical:
                            internal_fm = row_2[col_idx + 1] if (col_idx + 1) < len(row_2) else None
                            if internal_fm is not None:
                                try:
                                    val = int(float(internal_fm))
                                    if ms.practical_full_marks != val:
                                        ms.practical_full_marks = val
                                        changed = True
                                except Exception:
                                    pass
                                    
                        if changed:
                            ms.save()
                    except Exception:
                        pass
                    col_idx += 2
                
                # 2. Parse class-wide total attendance days from the Present Days column in Row 2
                if col_idx < len(row_2) and row_2[col_idx] is not None:
                    try:
                        class_total_days = int(float(row_2[col_idx]))
                    except Exception:
                        pass
                        
                student_start_row = 3
            else:
                student_start_row = 2

            saved = 0
            errors = []
            from django.core.exceptions import ValidationError
            for row_num, row in enumerate(ws.iter_rows(min_row=student_start_row, values_only=True), start=student_start_row):
                if not row[0]:
                    continue
                roll = str(row[0]).strip()
                student = students.get(roll)
                if not student:
                    continue

                col_idx = 2
                subject_marks_list = []
                try:
                    for subject in subjects:
                        try:
                            ms = subject
                        except Exception:
                            col_idx += 2
                            continue
                        theory_raw = row[col_idx] if col_idx < len(row) else None
                        internal_raw = row[col_idx + 1] if ms.has_practical and (col_idx + 1) < len(row) else None
                        col_idx += 2

                        # Parse theory marks / special value
                        theory_obtained = None
                        special_value = None
                        if theory_raw is not None:
                            val_str = str(theory_raw).strip().upper()
                            if val_str in ('AB', 'ABSENT'):
                                special_value = 'AB'
                            elif val_str in ('WH', 'WITHHELD'):
                                special_value = 'WH'
                            elif val_str in ('EX', 'EXEMPTED'):
                                special_value = 'EX'
                            elif val_str in ('', 'NONE', 'NULL'):
                                theory_obtained = None
                            else:
                                try:
                                    from decimal import Decimal
                                    theory_obtained = Decimal(str(theory_raw))
                                except Exception:
                                    raise ValidationError(f"Invalid theory marks format '{theory_raw}' for subject '{subject.name}'.")

                        # Parse internal marks / special value
                        internal_obtained = None
                        if internal_raw is not None and ms.has_practical:
                            val_str = str(internal_raw).strip().upper()
                            if val_str in ('AB', 'ABSENT'):
                                special_value = 'AB'
                            elif val_str in ('WH', 'WITHHELD'):
                                special_value = 'WH'
                            elif val_str in ('EX', 'EXEMPTED'):
                                special_value = 'EX'
                            elif val_str in ('', 'NONE', 'NULL'):
                                internal_obtained = None
                            else:
                                try:
                                    from decimal import Decimal
                                    internal_obtained = Decimal(str(internal_raw))
                                except Exception:
                                    raise ValidationError(f"Invalid internal marks format '{internal_raw}' for subject '{subject.name}'.")

                        subject_marks_list.append((subject, theory_obtained, internal_obtained, special_value))

                    # Parse attendance columns (Present Days)
                    present_days = None
                    total_days = class_total_days
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            present_days = int(float(row[col_idx]))
                        except Exception:
                            pass

                    with transaction.atomic():
                        for subject, theory_obtained, internal_obtained, special_value in subject_marks_list:
                            if subject.subject_type == Subject.SubjectType.OPTIONAL:
                                if (student.id, subject.id) not in optional_enrollments:
                                    continue

                            defaults = {
                                'school': school,
                                'special_value': special_value,
                                'theory_obtained': theory_obtained,
                                'internal_obtained': internal_obtained,
                                'entered_by': request.user,
                            }
                            if not getattr(request.user, 'is_teacher', False):
                                defaults['present_days'] = present_days
                                defaults['total_days'] = total_days

                            # Run model level validation via clean()
                            me = MarkEntry(
                                exam=exam, student=student, subject=subject,
                                **defaults
                            )
                            me.clean()

                            MarkEntry.objects.update_or_create(
                                exam=exam, student=student, subject=subject,
                                defaults=defaults
                            )
                            saved += 1
                except ValidationError as ve:
                    errors.append(f"Row {row_num} (Student {student.name}): {', '.join(ve.messages) if hasattr(ve, 'messages') else str(ve)}")
                except Exception as e:
                    errors.append(f"Row {row_num} (Student {student.name}): {str(e)}")

            if saved > 0:
                messages.success(request, f'{saved} mark entries imported successfully.')
            if errors:
                for err in errors[:10]:
                    messages.error(request, err)
                if len(errors) > 10:
                    messages.warning(request, f"And {len(errors) - 10} more validation errors.")
        except Exception as e:
            logger.exception('Bulk mark import failed for exam=%s class=%s', exam_id, class_id)
            messages.error(request, f'Import failed: {str(e)}. Please check the file format and try again.')

        return redirect('mark_entry', exam_id=exam_id, class_id=class_id)

    return render(request, 'marks/bulk_import.html', {
        'exam': exam, 'class_obj': cls
    })


@login_required
def mark_entry_template(request, exam_id, class_id):
    """Download Excel template for mark entry."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    school = request.user.school
    from apps.exams.models import Exam
    from apps.classes.models import Class
    from apps.students.models import Student
    from apps.subjects.models import Subject

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    cls = get_object_or_404(Class, pk=class_id, school=school)
    students = Student.objects.filter(class_obj=cls, school=school, is_active=True)
    subjects = Subject.objects.filter(class_obj=cls, school=school).select_related(
        ).order_by('order')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{exam.name[:20]} Marks'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    sub_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

    # Build headers
    headers = ['Roll No', 'Student Name']
    for subj in subjects:
        try:
            ms = subj
            headers.append(f'{subj.name}\nTheory')
            if ms.has_practical:
                headers.append(f'{subj.name}\nInternal')
            else:
                headers.append(f'{subj.name}\n(No Internal)')
        except Exception:
            headers.append(subj.name)
            headers.append('')

    headers.append('Present Days')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill if col <= 2 else sub_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    ws.row_dimensions[1].height = 40

    # Load existing mark entries to pre-populate
    from apps.marks.models import MarkEntry
    existing_marks = {}
    for me in MarkEntry.objects.filter(exam=exam, school=school,
                                        student__in=students, subject__in=subjects):
        existing_marks[(me.student_id, me.subject_id)] = me

    from apps.subjects.models import StudentSubjectEnrollment
    optional_enrollments = set(StudentSubjectEnrollment.objects.filter(
        student__in=students, subject__in=subjects
    ).values_list('student_id', 'subject_id'))
    
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Row 2: Full Marks / Total Attendance Row
    ws.cell(row=2, column=1, value="")
    cell_fm = ws.cell(row=2, column=2, value="Full Marks / Total")
    cell_fm.font = Font(bold=True)
    cell_fm.alignment = Alignment(horizontal='right')

    col_idx = 3
    for subj in subjects:
        try:
            ms = subj
            ws.cell(row=2, column=col_idx, value=ms.theory_full_marks)
            if ms.has_practical:
                ws.cell(row=2, column=col_idx + 1, value=ms.practical_full_marks)
        except Exception:
            pass
        col_idx += 2

    # Populate class-wide total attendance days (if any student has one saved)
    class_total_days = None
    for student in students:
        for subj in subjects:
            me = existing_marks.get((student.id, subj.id))
            if me and me.total_days is not None:
                class_total_days = me.total_days
                break
        if class_total_days is not None:
            break

    if class_total_days is not None:
        ws.cell(row=2, column=col_idx, value=class_total_days)

    # Populate Student Rows (Row 3 onwards)
    for row_num, student in enumerate(students, 3):
        ws.cell(row=row_num, column=1, value=student.roll_number)
        ws.cell(row=row_num, column=2, value=student.name)
        
        col_idx = 3
        present_days = None
        for subj in subjects:
            is_unmapped_optional = subj.subject_type == Subject.SubjectType.OPTIONAL and (student.id, subj.id) not in optional_enrollments
            
            if is_unmapped_optional:
                ws.cell(row=row_num, column=col_idx).fill = black_fill
                if subj and subj.has_practical:
                    ws.cell(row=row_num, column=col_idx + 1).fill = black_fill
            else:
                me = existing_marks.get((student.id, subj.id))
                if me:
                    if me.present_days is not None:
                        present_days = me.present_days
                    
                    # Pre-populate marks
                    if me.special_value:
                        ws.cell(row=row_num, column=col_idx, value=me.special_value)
                        if subj and subj.has_practical:
                            ws.cell(row=row_num, column=col_idx + 1, value=me.special_value)
                    else:
                        if me.theory_obtained is not None:
                            ws.cell(row=row_num, column=col_idx, value=float(me.theory_obtained))
                        if subj and subj.has_practical and me.internal_obtained is not None:
                            ws.cell(row=row_num, column=col_idx + 1, value=float(me.internal_obtained))
            col_idx += 2
            
        # Write present days column at the end
        if present_days is not None:
            ws.cell(row=row_num, column=col_idx, value=present_days)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    clean_class = cls.full_name.replace(' ', '_')
    clean_exam = exam.name.replace('Examination', '').replace('exam', '').strip().replace(' ', '_')
    response['Content-Disposition'] = (
        f'attachment; filename="ME_{clean_class}_{clean_exam}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def save_full_marks(request):
    """AJAX endpoint to update subject full marks."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    school = request.user.school
    from apps.subjects.models import Subject

    subject = get_object_or_404(Subject, pk=data.get('subject_id'), school=school)
    field = data.get('field')  # 'theory' or 'internal'

    # SECURITY: Validate bounds to prevent mark manipulation via fraudulent full marks
    try:
        value = int(data.get('value') or 100)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid value. Must be an integer.'}, status=400)

    MIN_FULL_MARKS = 1
    MAX_FULL_MARKS = 1000
    if not (MIN_FULL_MARKS <= value <= MAX_FULL_MARKS):
        return JsonResponse(
            {'error': f'Full marks must be between {MIN_FULL_MARKS} and {MAX_FULL_MARKS}.'},
            status=400
        )

    try:
        if field == 'theory':
            subject.theory_full_marks = value
        elif field == 'internal':
            subject.practical_full_marks = value
        subject.save()
        return JsonResponse({'status': 'saved'})
    except Exception:
        logger.exception('Error in save_full_marks for subject=%s field=%s', data.get('subject_id'), field)
        return JsonResponse({'error': 'An unexpected error occurred.'}, status=500)


@login_required
def mark_entry_select(request):
    """Mark Entry selection page - choose an exam and class to proceed."""
    import json
    school = request.user.school
    from apps.exams.models import Exam, ExamClass
    from apps.classes.models import Class
    
    active_session = school.get_active_session() if school else None

    # Fetch all exams
    exams = Exam.objects.filter(school=school)
    if active_session:
        exams = exams.filter(session=active_session)
    exams = exams.select_related('session').order_by('-created_at')
    # Fetch classes with consistent ordering
    classes = Class.objects.filter(school=school).order_by('numeric_level', 'name', 'section')

    # Build mapping of exam ID to its associated classes
    exam_classes_map = {}
    exam_classes = ExamClass.objects.filter(exam__school=school)
    if active_session:
        exam_classes = exam_classes.filter(exam__session=active_session)
    exam_classes = exam_classes.select_related('class_obj')
    for ec in exam_classes:
        exam_classes_map.setdefault(ec.exam_id, []).append({
            'id': ec.class_obj.id,
            'name': ec.class_obj.full_name
        })

    exam_classes_json = json.dumps(exam_classes_map)

    # Check for selected parameters
    exam_id = request.GET.get('exam')
    class_id = request.GET.get('class_obj')
    
    if exam_id and class_id:
        return redirect('mark_entry', exam_id=exam_id, class_id=class_id)

    context = {
        'exams': exams,
        'exam_classes_json': exam_classes_json,
        'selected_exam_id': int(exam_id) if exam_id and exam_id.isdigit() else None,
        'school': school,
    }

    return render(request, 'marks/entry_select.html', context)

@login_required
def exam_progress_ajax(request, exam_id):
    from apps.exams.models import Exam
    from apps.marks.models import MarkEntry
    from apps.students.models import Student
    from apps.subjects.models import Subject
    from django.http import JsonResponse
    
    exam = get_object_or_404(Exam, pk=exam_id, school=request.user.school)
    
    exam_classes = exam.exam_classes.select_related('class_obj').all()
    
    total_expected = 0
    for ec in exam_classes:
        cls = ec.class_obj
        num_students = Student.objects.filter(class_obj=cls, is_active=True).count()
        num_subjects = Subject.objects.filter(class_obj=cls).count()
        total_expected += (num_students * num_subjects)
        
    total_entered = MarkEntry.objects.filter(exam=exam).count()
    
    percentage = 0
    if total_expected > 0:
        percentage = round((total_entered / total_expected) * 100, 1)
        if percentage > 100: percentage = 100
        
    return JsonResponse({
        'exam_id': exam.id,
        'total_expected': total_expected,
        'total_entered': total_entered,
        'percentage': percentage
    })


@login_required
@require_POST
def toggle_class_lock(request, exam_id, class_id):
    """Admin toggles the lock state of a class for an exam, processing results if locked."""
    if not request.user.can_manage_school():
        messages.error(request, 'Only administrators can lock/unlock results.')
        return redirect('mark_entry', exam_id=exam_id, class_id=class_id)
        
    school = request.user.school
    from apps.exams.models import ExamClass
    from apps.results.services import ResultProcessingService
    
    exam_class = get_object_or_404(ExamClass, exam_id=exam_id, class_obj_id=class_id, exam__school=school)
    
    action = request.POST.get('action')
    if action == 'lock':
        exam_class.is_locked = True
        exam_class.save()
        # Auto process results
        service = ResultProcessingService(exam_class.exam)
        service.process(class_obj=exam_class.class_obj)
        messages.success(request, f'Results for {exam_class.class_obj.name} have been processed and locked.')
    elif action == 'unlock':
        exam_class.is_locked = False
        exam_class.save()
        messages.success(request, f'Results for {exam_class.class_obj.name} are unlocked and editable.')
        
    return redirect('mark_entry', exam_id=exam_id, class_id=class_id)


@login_required
def mark_entry_all_template(request, exam_id):
    """Download a single Excel file containing templates for all classes as sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    school = request.user.school
    from apps.exams.models import Exam, ExamClass
    from apps.students.models import Student
    from apps.subjects.models import Subject
    from apps.marks.models import MarkEntry

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    exam_classes = ExamClass.objects.filter(exam=exam, exam__school=school).select_related('class_obj')

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    sub_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

    for ec in exam_classes:
        cls = ec.class_obj
        students = Student.objects.filter(class_obj=cls, school=school, is_active=True)
        if not students.exists():
            continue

        subjects = Subject.objects.filter(class_obj=cls, school=school).select_related(
            ).order_by('order')

        # Sheet names must be <= 31 chars
        sheet_title = f"{cls.full_name}"[:31].replace(':', '-').replace('/', '-')
        ws = wb.create_sheet(title=sheet_title)

        # Build headers
        headers = ['Roll No', 'Student Name']
        for subj in subjects:
            try:
                ms = subj
                headers.append(f'{subj.name}\nTheory')
                if ms.has_practical:
                    headers.append(f'{subj.name}\nInternal')
                else:
                    headers.append(f'{subj.name}\n(No Internal)')
            except Exception:
                headers.append(subj.name)
                headers.append('')

        headers.append('Present Days')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill if col <= 2 else sub_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        ws.row_dimensions[1].height = 40

        # Load existing mark entries to pre-populate
        existing_marks = {}
        for me in MarkEntry.objects.filter(exam=exam, school=school,
                                            student__in=students, subject__in=subjects):
            existing_marks[(me.student_id, me.subject_id)] = me

        # Row 2: Full Marks / Total Attendance Row
        ws.cell(row=2, column=1, value="")
        cell_fm = ws.cell(row=2, column=2, value="Full Marks / Total")
        cell_fm.font = Font(bold=True)
        cell_fm.alignment = Alignment(horizontal='right')

        col_idx = 3
        for subj in subjects:
            try:
                ms = subj
                ws.cell(row=2, column=col_idx, value=ms.theory_full_marks)
                if ms.has_practical:
                    ws.cell(row=2, column=col_idx + 1, value=ms.practical_full_marks)
            except Exception:
                pass
            col_idx += 2

        # Populate class-wide total attendance days
        class_total_days = None
        for student in students:
            for subj in subjects:
                me = existing_marks.get((student.id, subj.id))
                if me and me.total_days is not None:
                    class_total_days = me.total_days
                    break
            if class_total_days is not None:
                break

        if class_total_days is not None:
            ws.cell(row=2, column=col_idx, value=class_total_days)

        # Populate Student Rows (Row 3 onwards)
        for row_num, student in enumerate(students, 3):
            ws.cell(row=row_num, column=1, value=student.roll_number)
            ws.cell(row=row_num, column=2, value=student.name)
            
            col_idx = 3
            present_days = None
            for subj in subjects:
                me = existing_marks.get((student.id, subj.id))
                if me:
                    if me.present_days is not None:
                        present_days = me.present_days
                    
                    # Pre-populate marks
                    if me.special_value:
                        ws.cell(row=row_num, column=col_idx, value=me.special_value)
                        if subj and subj.has_practical:
                            ws.cell(row=row_num, column=col_idx + 1, value=me.special_value)
                    else:
                        if me.theory_obtained is not None:
                            ws.cell(row=row_num, column=col_idx, value=float(me.theory_obtained))
                        if subj and subj.has_practical and me.internal_obtained is not None:
                            ws.cell(row=row_num, column=col_idx + 1, value=float(me.internal_obtained))
                col_idx += 2
                
            # Write present days column at the end
            if present_days is not None:
                ws.cell(row=row_num, column=col_idx, value=present_days)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    clean_exam = exam.name.replace('Examination', '').replace('exam', '').strip().replace(' ', '_')
    response['Content-Disposition'] = (
        f'attachment; filename="ME_All_Classes_{clean_exam}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@require_POST
def bulk_mark_all_import(request, exam_id):
    """Import marks/attendance for all classes from a multi-sheet Excel file."""
    school = request.user.school
    from apps.exams.models import Exam, ExamClass
    from apps.classes.models import Class
    from apps.students.models import Student
    from apps.subjects.models import Subject
    from apps.marks.models import MarkEntry

    exam = get_object_or_404(Exam, pk=exam_id, school=school)
    if not request.FILES.get('excel_file'):
        messages.error(request, 'Please upload an Excel file.')
        return redirect('mark_entry_select')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(request.FILES['excel_file'])
        
        saved_total = 0
        skipped_sheets = []
        errors = []
        
        # Get all class names mapping to class objects
        classes_map = {}
        for c in Class.objects.filter(school=school):
            classes_map[c.full_name.lower().strip()] = c
            classes_map[c.name.lower().strip()] = c

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cls = classes_map.get(sheet_name.lower().strip())
            if not cls:
                skipped_sheets.append(sheet_name)
                continue

            exam_class = ExamClass.objects.filter(exam=exam, class_obj=cls).first()
            if not exam_class or exam_class.is_locked:
                continue

            subjects = list(Subject.objects.filter(class_obj=cls, school=school).order_by('order'))
            students = {s.roll_number: s for s in Student.objects.filter(class_obj=cls, school=school)}

            # Validate headers to prevent users from reordering columns
            expected_headers = ['Roll No', 'Student Name']
            for subj in subjects:
                try:
                    ms = subj
                    expected_headers.append(f'{subj.name}\nTheory')
                    if ms.has_practical:
                        expected_headers.append(f'{subj.name}\nInternal')
                    else:
                        expected_headers.append(f'{subj.name}\n(No Internal)')
                except Exception:
                    expected_headers.append(subj.name)
                    expected_headers.append('')
            expected_headers.append('Present Days')
            
            actual_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            actual_headers_clean = [str(h).strip() if h is not None else '' for h in actual_headers][:len(expected_headers)]
            expected_headers_clean = [str(h).strip() for h in expected_headers]
            
            if actual_headers_clean != expected_headers_clean:
                errors.append(f"Sheet '{sheet_name}': Column structure mismatch. Skipping.")
                continue

            # Check if Row 2 is the Full Marks / Total config row
            row_2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            has_config_row = False
            class_total_days = None
            
            if len(row_2) > 1 and row_2[1] and any(x in str(row_2[1]).lower() for x in ('full marks', 'total')):
                has_config_row = True
                
            if has_config_row:
                # SECURITY: Only allow School Admins to update the marking structure from Excel.
                # Teachers uploading Excel should NOT be able to manipulate full marks.
                is_admin_upload = request.user.is_school_admin or request.user.is_super_admin

                # 1. Parse and update subject full marks (admin only)
                col_idx = 2
                for subject in subjects:
                    try:
                        ms = subject
                        changed = False

                        theory_fm = row_2[col_idx] if col_idx < len(row_2) else None
                        if theory_fm is not None and is_admin_upload:
                            try:
                                val = int(float(theory_fm))
                                # SECURITY: Bounds check — prevent fraudulent marking structures
                                if 1 <= val <= 1000 and ms.theory_full_marks != val:
                                    ms.theory_full_marks = val
                                    changed = True
                            except Exception:
                                pass

                        if ms.has_practical:
                            internal_fm = row_2[col_idx + 1] if (col_idx + 1) < len(row_2) else None
                            if internal_fm is not None and is_admin_upload:
                                try:
                                    val = int(float(internal_fm))
                                    if 1 <= val <= 1000 and ms.practical_full_marks != val:
                                        ms.practical_full_marks = val
                                        changed = True
                                except Exception:
                                    pass

                        if changed:
                            ms.save()
                    except Exception:
                        pass
                    col_idx += 2
                
                # 2. Parse class-wide total attendance days from the Present Days column in Row 2
                if col_idx < len(row_2) and row_2[col_idx] is not None:
                    try:
                        class_total_days = int(float(row_2[col_idx]))
                    except Exception:
                        pass
                        
                student_start_row = 3
            else:
                student_start_row = 2

            for row_num, row in enumerate(ws.iter_rows(min_row=student_start_row, values_only=True), start=student_start_row):
                try:
                    if not row[0]:
                        continue
                    roll = str(row[0]).strip()
                    student = students.get(roll)
                    if not student:
                        continue
    
                    col_idx = 2
                    subject_marks_list = []
                    for subject in subjects:
                        try:
                            ms = subject
                        except Exception:
                            col_idx += 2
                            continue
                        theory_raw = row[col_idx] if col_idx < len(row) else None
                        internal_raw = row[col_idx + 1] if ms.has_practical and (col_idx + 1) < len(row) else None
                        col_idx += 2
    
                        # Parse theory marks / special value
                        theory_obtained = None
                        special_value = None
                        if theory_raw is not None:
                            val_str = str(theory_raw).strip().upper()
                            if val_str in ('AB', 'ABSENT'):
                                special_value = 'AB'
                            elif val_str in ('WH', 'WITHHELD'):
                                special_value = 'WH'
                            elif val_str in ('EX', 'EXEMPTED'):
                                special_value = 'EX'
                            elif val_str in ('', 'NONE', 'NULL'):
                                theory_obtained = None
                            else:
                                try:
                                    from decimal import Decimal
                                    theory_obtained = Decimal(str(theory_raw))
                                except Exception:
                                    raise ValueError(f"Invalid theory marks format '{theory_raw}' in sheet '{sheet_name}', row {row_num}, for subject '{subject.name}'.")
    
                        # Parse internal marks / special value
                        internal_obtained = None
                        if internal_raw is not None and ms.has_practical:
                            val_str = str(internal_raw).strip().upper()
                            if val_str in ('AB', 'ABSENT'):
                                special_value = 'AB'
                            elif val_str in ('WH', 'WITHHELD'):
                                special_value = 'WH'
                            elif val_str in ('EX', 'EXEMPTED'):
                                special_value = 'EX'
                            elif val_str in ('', 'NONE', 'NULL'):
                                internal_obtained = None
                            else:
                                try:
                                    from decimal import Decimal
                                    internal_obtained = Decimal(str(internal_raw))
                                except Exception:
                                    raise ValueError(f"Invalid internal marks format '{internal_raw}' in sheet '{sheet_name}', row {row_num}, for subject '{subject.name}'.")
    
                        subject_marks_list.append((subject, theory_obtained, internal_obtained, special_value))
    
                    # Parse attendance (Present Days)
                    present_days = None
                    total_days = class_total_days
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            present_days = int(float(row[col_idx]))
                        except Exception:
                            pass
    
                    from django.core.exceptions import ValidationError
                    for subject, theory_obtained, internal_obtained, special_value in subject_marks_list:
                        defaults = {
                            'school': school,
                            'special_value': special_value,
                            'theory_obtained': theory_obtained,
                            'internal_obtained': internal_obtained,
                            'entered_by': request.user,
                        }
                        if not getattr(request.user, 'is_teacher', False):
                            defaults['present_days'] = present_days
                            defaults['total_days'] = total_days
    
                        # Run model level validation via clean()
                        me = MarkEntry(
                            exam=exam, student=student, subject=subject,
                            **defaults
                        )
                        try:
                            me.clean()
                        except ValidationError as ve:
                            raise ValueError(f"Sheet '{sheet_name}', Row {row_num} (Student {roll}): {', '.join(ve.messages) if hasattr(ve, 'messages') else str(ve)}")
    
                        MarkEntry.objects.update_or_create(
                            exam=exam, student=student, subject=subject,
                            defaults=defaults
                        )
                        saved_total += 1
                except ValueError as ve:
                    errors.append(str(ve))
                except Exception as e:
                    errors.append(f"Sheet '{sheet_name}', Row {row_num} (Student {roll}): {str(e)}")

        if saved_total > 0:
            msg = f'Successfully imported {saved_total} mark entries across classes.'
            if skipped_sheets:
                msg += f" Skipped unknown sheets: {', '.join(skipped_sheets)}."
            messages.success(request, msg)

        if errors:
            for err in errors[:10]:
                messages.error(request, err)
            if len(errors) > 10:
                messages.warning(request, f"And {len(errors) - 10} more validation errors.")

    except Exception as e:
        logger.exception('Bulk all-class import failed for exam=%s', exam_id)
        messages.error(request, f'Bulk import failed: {str(e)}. Please check the file format and try again.')

    return redirect('mark_entry_select')
