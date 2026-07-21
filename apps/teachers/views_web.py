"""
Teachers App — Web views
"""
import io
import json
import openpyxl
import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db import models as db_models

from apps.accounts.models import User
from apps.classes.models import Class
from apps.subjects.models import Subject
from .models import Teacher, TeacherSubject


@login_required
def teacher_list(request):
    """List all teachers in the school."""
    school = request.user.school
    
    if request.method == 'POST':
        if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
            messages.error(request, 'Access denied.')
            return redirect('teacher_list')

    active_session = school.get_active_session()
    
    teachers_qs = Teacher.objects.filter(school=school).select_related('user').prefetch_related('assigned_classes', 'subject_assignments__subject')
    if active_session:
        teachers_qs = teachers_qs.filter(sessions=active_session)
    
    def get_sort_key(t):
        try:
            val = float(t.sn) if t.sn else float('inf')
        except ValueError:
            import re
            nums = re.findall(r'\d+', t.sn)
            val = float(nums[0]) if nums else float('inf')
        return (val, t.name)
        
    teachers = sorted(teachers_qs, key=get_sort_key)
    
    return render(request, 'teachers/list.html', {
        'teachers': teachers
    })


@login_required
def teacher_create(request):
    """Create a new teacher manually."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    if request.method == 'POST':
        sn = request.POST.get('sn', '').strip()
        name = request.POST.get('name', '').strip()
        contact = request.POST.get('contact_number', '').strip()
        dob = request.POST.get('date_of_birth') or None
        email = request.POST.get('email', '').strip() or None

        if email and Teacher.objects.filter(email__iexact=email).exists():
            messages.error(request, f'The email {email} is already in use by another teacher.')
            return render(request, 'teachers/form.html')

        photo = None
        if 'photo' in request.FILES:
            from django.core.exceptions import ValidationError
            from core.security import validate_image_upload
            try:
                validate_image_upload(request.FILES['photo'])
                photo = request.FILES['photo']
            except ValidationError as e:
                messages.error(request, str(e.message))
                return render(request, 'teachers/form.html')

        teacher = Teacher.objects.create(
            school=school,
            sn=sn,
            name=name,
            contact_number=contact,
            date_of_birth=dob,
            email=email,
            photo=photo
        )
        active_session = school.get_active_session()
        if active_session:
            teacher.sessions.add(active_session)
            
        messages.success(request, f'Teacher {name} added successfully.')
        return redirect('teacher_list')

    return render(request, 'teachers/form.html')


@login_required
def teacher_detail(request, pk):
    """View teacher profile."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    teacher = get_object_or_404(
        Teacher.objects.prefetch_related('assigned_classes', 'subject_assignments__subject'),
        pk=pk, 
        school=school
    )
    
    # Get active session assignments
    active_session = school.get_active_session() if school else None
    assignments = teacher.assigned_classes.all()
    subject_assignments = teacher.subject_assignments.all()
    if active_session:
        assignments = assignments.filter(session=active_session)
        
    return render(request, 'teachers/detail.html', {
        'teacher': teacher,
        'assignments': assignments,
        'subject_assignments': subject_assignments,
        'active_session': active_session
    })


@login_required
def teacher_edit(request, pk):
    """Edit teacher details."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    teacher = get_object_or_404(Teacher, pk=pk, school=school)

    if request.method == 'POST':
        teacher.sn = request.POST.get('sn', teacher.sn).strip()
        teacher.name = request.POST.get('name', teacher.name).strip()
        teacher.contact_number = request.POST.get('contact_number', teacher.contact_number).strip()
        dob = request.POST.get('date_of_birth')
        teacher.date_of_birth = dob if dob else None
        
        email = request.POST.get('email', '').strip() or None
        if email and email != teacher.email:
            if Teacher.objects.exclude(pk=teacher.pk).filter(email__iexact=email).exists():
                messages.error(request, f'The email {email} is already in use by another teacher.')
                return render(request, 'teachers/form.html', {'teacher': teacher})
        teacher.email = email
        if teacher.user and teacher.user.email != (email or ''):
            teacher.user.email = email or ''
            teacher.user.save()

        if 'photo' in request.FILES:
            from django.core.exceptions import ValidationError
            from core.security import validate_image_upload
            try:
                validate_image_upload(request.FILES['photo'])
                teacher.photo = request.FILES['photo']
            except ValidationError as e:
                messages.error(request, str(e.message))
                return render(request, 'teachers/form.html', {'teacher': teacher})
        elif request.POST.get('remove_photo'):
            if teacher.photo:
                teacher.photo.delete(save=False)
                teacher.photo = None

        teacher.save()
        messages.success(request, f'Teacher {teacher.name} updated successfully.')
        return redirect('teacher_list')

    return render(request, 'teachers/form.html', {'teacher': teacher})


@login_required
def teacher_delete(request, pk):
    """Delete a teacher."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    
    if request.method == 'POST':
        name = teacher.name
        # Delete associated user account if it exists
        if teacher.user:
            teacher.user.delete()
        teacher.delete()
        messages.success(request, f'Teacher {name} permanently deleted.')
        return redirect('teacher_list')
        
    return redirect('teacher_list')


@login_required
def teacher_bulk_delete(request):
    """Bulk delete teachers."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    if request.method == 'POST':
        teacher_ids = request.POST.getlist('teacher_ids')
        if teacher_ids:
            teachers = Teacher.objects.filter(pk__in=teacher_ids, school=school)
            deleted_count = 0
            for teacher in teachers:
                if teacher.user:
                    teacher.user.delete()
                teacher.delete()
                deleted_count += 1
            if deleted_count > 0:
                messages.success(request, f'Successfully deleted {deleted_count} teacher(s).')
            else:
                messages.warning(request, 'No teachers were deleted.')
        else:
            messages.warning(request, 'No teachers were selected for deletion.')
            
    return redirect('teacher_list')


@login_required
def teacher_import(request):
    """Import teachers from Excel."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            created = 0
            errors = []
            
            from django.db import transaction
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0] and not row[1]:
                        continue
                    try:
                        sn = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                        if not name:
                            continue
                            
                        contact = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                        
                        dob = None
                        dob_val = row[3] if len(row) > 3 else None
                        if dob_val is not None:
                            if isinstance(dob_val, (datetime.date, datetime.datetime)):
                                dob = dob_val if isinstance(dob_val, datetime.date) else dob_val.date()
                            else:
                                dob_str = str(dob_val).strip()
                                if dob_str:
                                    date_formats = ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d')
                                    for fmt in date_formats:
                                        try:
                                            dob = datetime.datetime.strptime(dob_str, fmt).date()
                                            break
                                        except ValueError:
                                            continue

                        teacher = Teacher.objects.create(
                            school=school,
                            sn=sn,
                            name=name,
                            contact_number=contact,
                            date_of_birth=dob
                        )
                        active_session = school.get_active_session()
                        if active_session:
                            teacher.sessions.add(active_session)
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {row_num}: {str(e)}')
                        
                if errors:
                    transaction.set_rollback(True)
                    messages.error(request, f'Import failed due to errors. No teachers were imported.')
                    for err in errors[:5]:
                        messages.error(request, err)
                    if len(errors) > 5:
                        messages.error(request, f'...and {len(errors) - 5} more errors.')
                else:
                    messages.success(request, f'{created} teachers imported successfully.')

        except Exception as e:
            messages.error(request, f'Failed to read Excel file: {str(e)}')

        return redirect('teacher_list')

    return render(request, 'teachers/import.html')


@login_required
def teacher_import_template(request):
    """Download blank Excel template for teacher import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Teachers'
    
    from openpyxl.styles import Font, PatternFill
    headers = ['SN', 'Teacher Name*', 'Contact Number', 'Date of Birth']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="teacher_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
def teacher_create_user(request, pk):
    """Auto-generate a User account for the teacher."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    
    if request.method == 'POST':
        if teacher.user:
            messages.warning(request, f'Teacher {teacher.name} already has an account.')
            return redirect('teacher_list')
            
        import re
        # Generate base username
        base_username = re.sub(r'[^a-zA-Z0-9]', '', teacher.name.lower())
        if teacher.sn:
            username = f"{base_username}{teacher.sn}"
        else:
            username = base_username
            
        # Ensure uniqueness
        original_username = username
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{original_username}{counter}"
            counter += 1
        import secrets
        # Generate an 8-character password
        password = secrets.token_urlsafe(6)

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    email=teacher.email or '',
                    password=password,
                    first_name=teacher.name,
                    role=User.Role.TEACHER,
                    school=school
                )
                teacher.user = user
                teacher.save()

            messages.success(request, f'Account created for {teacher.name}. Username: {username}. Temporary Password: {password} (Please copy this password now, it will not be shown again!)')
            # In a real system, you'd likely email it or show it ONCE on a dedicated screen, not in the session flash messages.
        except Exception as e:
            messages.error(request, f'Error creating account: {str(e)}')

        return redirect('teacher_list')

    return redirect('teacher_list')


@login_required
def teacher_reset_password(request, pk):
    """Reset password for a teacher's User account."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    
    if request.method == 'POST':
        if not teacher.user:
            messages.error(request, 'This teacher does not have an account.')
            return redirect('teacher_list')
        import secrets
        # Generate an 8-character password
        password = secrets.token_urlsafe(6)

        teacher.user.set_password(password)
        teacher.user.save()
        messages.success(request, f'Password reset for {teacher.name}. New password: {password} (Please copy this password now!)')
    return redirect('teacher_list')


@login_required
def teacher_send_password_reset(request, pk):
    """Send a password reset email to the teacher."""
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    
    if request.method == 'POST':
        if not teacher.email:
            messages.error(request, 'This teacher does not have an email address set.')
            return redirect(request.META.get('HTTP_REFERER', 'teacher_list'))
            
        import secrets
        from django.contrib.auth.tokens import default_token_generator
        from django.utils.http import urlsafe_base64_encode
        from django.utils.encoding import force_bytes
        from django.urls import reverse
        from django.core.mail import send_mail
        from django.conf import settings
        
        try:
            if not teacher.user:
                # Auto-create user so they can receive the reset link and log in
                username_base = teacher.email.split('@')[0]
                user = User.objects.create_user(
                    username=f"{username_base}_{teacher.pk}_{secrets.token_hex(2)}",
                    email=teacher.email,
                    password=secrets.token_urlsafe(16),
                    first_name=teacher.name,
                    role=User.Role.TEACHER,
                    school=school
                )
                teacher.user = user
                teacher.save()
            else:
                # Ensure email matches
                if teacher.user.email != teacher.email:
                    teacher.user.email = teacher.email
                    teacher.user.save()
                    
            uid = urlsafe_base64_encode(force_bytes(teacher.user.pk))
            token = default_token_generator.make_token(teacher.user)
            reset_url = request.build_absolute_uri(
                reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
            )
            
            from core.email_utils import teacher_password_reset_email
            _subject, _plain, _html = teacher_password_reset_email(teacher.name, reset_url)
            send_mail(
                subject=_subject,
                message=_plain,
                from_email=getattr(settings, 'EMAIL_HOST_USER', None) or "noreply@rms.local",
                recipient_list=[teacher.email],
                fail_silently=False,
                html_message=_html,
            )
            messages.success(request, f'Password reset email sent to {teacher.email}.')
            
            # Log action
            from apps.audit.models import AuditLog
            AuditLog.log_action(
                user=request.user,
                action='PASSWORD_RESET_SENT',
                model_name='Teacher',
                object_id=teacher.pk,
                details=f'Sent password reset email to {teacher.email}'
            )
        except Exception as e:
            messages.error(request, f'Error sending password reset email: {str(e)}')

    return redirect(request.META.get('HTTP_REFERER', 'teacher_list'))


# ---------------------------------------------------------
# TEACHER PORTAL VIEWS
# ---------------------------------------------------------

@login_required
def teacher_dashboard(request):
    """Dashboard specifically for users with the TEACHER role."""
    if not request.user.is_teacher:
        return redirect('dashboard')
        
    try:
        teacher = request.user.teacher_profile
    except Teacher.DoesNotExist:
        messages.error(request, 'Your account is not linked to any teacher profile.')
        return render(request, 'teachers/teacher_portal/dashboard.html', {'error': True})
        
    active_session = request.user.school.get_active_session() if hasattr(request.user, 'school') and request.user.school else None
    
    from apps.classes.models import Class
    from django.db.models import Q
    from apps.exams.models import ExamClass
    
    # 1. My Classes (Class Teacher)
    my_classes_qs = Class.objects.filter(class_teacher=teacher)
    if active_session:
        my_classes_qs = my_classes_qs.filter(session=active_session)
        
    my_classes_info = []
    for cls in my_classes_qs:
        subjects = cls.subjects.all()
        latest_ec = ExamClass.objects.filter(
            class_obj=cls, 
            exam__status='DRAFT'
        )
        if active_session:
            latest_ec = latest_ec.filter(exam__session=active_session)
        latest_ec = latest_ec.order_by('-exam__created_at').first()
        
        my_classes_info.append({
            'class_obj': cls,
            'subjects': subjects,
            'latest_exam_id': latest_ec.exam_id if latest_ec else None
        })

    # 2. My Subjects (Assigned Subjects)
    assigned_class_ids = teacher.subject_assignments.values_list('subject__class_obj_id', flat=True).distinct()
    assigned_classes_qs = Class.objects.filter(id__in=assigned_class_ids)
    if active_session:
        assigned_classes_qs = assigned_classes_qs.filter(session=active_session)
        
    my_subjects_info = []
    my_subject_list = []  # flat list: [{class_obj, subject, latest_exam_id}]
    for cls in assigned_classes_qs:
        assigned_subject_ids = teacher.subject_assignments.filter(subject__class_obj=cls).values_list('subject_id', flat=True)
        subjects = cls.subjects.filter(id__in=assigned_subject_ids)
        
        latest_ec = ExamClass.objects.filter(
            class_obj=cls, 
            exam__status='DRAFT'
        )
        if active_session:
            latest_ec = latest_ec.filter(exam__session=active_session)
        latest_ec = latest_ec.order_by('-exam__created_at').first()
        
        my_subjects_info.append({
            'class_obj': cls,
            'subjects': subjects,
            'latest_exam_id': latest_ec.exam_id if latest_ec else None
        })
        for sub in subjects:
            my_subject_list.append({
                'class_obj': cls,
                'subject': sub,
                'latest_exam_id': latest_ec.exam_id if latest_ec else None
            })

    classes_qs = Class.objects.filter(
        Q(id__in=my_classes_qs.values_list('id')) | Q(id__in=assigned_classes_qs.values_list('id'))
    ).distinct()

        
    # Calculate KPIs
    from apps.students.models import Student
    from apps.exams.models import Exam
    from apps.results.models import StudentResult
    from django.db.models import Avg

    assignments = teacher.subject_assignments.all()
    if active_session:
        assignments = assignments.filter(subject__session=active_session)
    class_ids = classes_qs.values_list('id', flat=True)

    total_classes = classes_qs.count()
    total_subjects = assignments.count()
    primary_subject = assignments.first().subject.name if assignments.exists() else "None"
    
    total_students = Student.objects.filter(class_obj__id__in=class_ids, is_active=True).count()
    
    # Exams related to the active session for these classes
    related_exams = Exam.objects.filter(session=active_session, exam_classes__class_obj__id__in=class_ids).distinct()
    
    # Published vs Pending logic
    published_exams = related_exams.filter(status='PUBLISHED').count()
    total_exams = related_exams.count()
    published_percentage = int((published_exams / total_exams * 100)) if total_exams > 0 else 0
    
    pending_marks = total_exams - published_exams  # Simplified pending logic
    
    # Get the single most recent published exam for global KPIs (or most recent overall)
    recent_exam = related_exams.filter(status='PUBLISHED').order_by('-start_date', '-created_at').first()
    if not recent_exam:
        recent_exam = related_exams.order_by('-start_date', '-created_at').first()
    
    # Average GPA
    if recent_exam:
        avg_gpa_dict = StudentResult.objects.filter(exam=recent_exam, student__class_obj__id__in=class_ids).aggregate(Avg('overall_gpa'))
        avg_gpa = round(avg_gpa_dict['overall_gpa__avg'] or 0.0, 2)
    else:
        avg_gpa = 0.0
    
    # Grade Distribution
    from django.db.models import Count
    import json
    if recent_exam:
        grade_counts = StudentResult.objects.filter(exam=recent_exam, student__class_obj__id__in=class_ids).values('final_grade').annotate(count=Count('id')).order_by('final_grade')
    else:
        grade_counts = []
    
    grades = []
    counts = []
    for gc in grade_counts:
        grade = gc['final_grade']
        if grade == '':
            continue
        grades.append(grade)
        counts.append(gc['count'])
        
    grade_labels_json = json.dumps(grades)
    grade_data_json = json.dumps(counts)

    # Top Performers per Class and Subject
    from django.db.models.functions import Coalesce
    from django.db.models import DecimalField
    from apps.marks.models import MarkEntry
    
    top_performers = []
    # Only show top performers for subjects actually taught by this teacher
    for item in my_subject_list:
        cls_obj = item['class_obj']
        subject = item['subject']
        
        top_marks_qs = MarkEntry.objects.filter(
            student__class_obj=cls_obj, 
            subject=subject,
            session=active_session
        )
        if recent_exam:
            top_marks_qs = top_marks_qs.filter(exam=recent_exam)
            
        top_marks = top_marks_qs.annotate(
            total_marks=Coalesce('theory_obtained', 0.0, output_field=DecimalField()) + Coalesce('internal_obtained', 0.0, output_field=DecimalField())
        ).order_by('-total_marks')[:3]

        leaders = []
        for idx, mark in enumerate(top_marks):
            leaders.append({
                'rank': idx + 1,
                'student': mark.student,
                'score': float(mark.total_marks) if mark.total_marks else 0.0
            })
        
        if leaders:
            top_performers.append({
                'class': cls_obj,
                'subject': subject,
                'leaders': leaders
            })
    
    # Sort top performers by class level, then subject
    top_performers.sort(key=lambda x: (x['class'].numeric_level, x['class'].name, x['subject'].name))

    # ── NEW CLASS TEACHER ANALYTICS ──
    from apps.marks.models import MarkEntry
    from django.db.models import Avg, DecimalField
    from django.db.models.functions import Coalesce

    # Only compute detailed analytics for class teacher classes
    class_analytics = []
    for info in my_classes_info:
        cls_obj = info['class_obj']
        cls_students = list(Student.objects.filter(class_obj=cls_obj, school=request.user.school, is_active=True).order_by('roll_number'))
        cls_subjects = list(cls_obj.subjects.all().order_by('order'))

        # --- Subject-wise average ---
        subject_averages = []
        for subject in cls_subjects:
            avg_qs = MarkEntry.objects.filter(
                student__class_obj=cls_obj,
                subject=subject,
                session=active_session
            )
            if recent_exam:
                avg_qs = avg_qs.filter(exam=recent_exam)
            
            avg = avg_qs.aggregate(
                th_avg=Avg('theory_obtained'),
                in_avg=Avg('internal_obtained'),
            )
            th_avg = float(avg['th_avg'] or 0)
            in_avg = float(avg['in_avg'] or 0)
            total_avg = th_avg + in_avg
            full = float((subject.theory_full_marks or 0) + (subject.practical_full_marks or 0))
            pct = round((total_avg / full * 100), 1) if full > 0 else 0
            subject_averages.append({
                'subject': subject,
                'avg': round(total_avg, 1),
                'pct': pct,
                'full': full,
            })

        # --- Top 3 students by GPA ---
        # --- Top 3/5 students by GPA ---
        # Find the latest published exam for this class
        from apps.exams.models import ExamClass as EC
        latest_ec_qs = EC.objects.filter(
            class_obj=cls_obj,
            exam__status='PUBLISHED',
        )
        if active_session:
            latest_ec_qs = latest_ec_qs.filter(exam__session=active_session)
        latest_ec_obj = latest_ec_qs.order_by('-exam__created_at').first()

        top_students = []
        if latest_ec_obj:
            results_qs = StudentResult.objects.filter(
                student__class_obj=cls_obj,
                exam=latest_ec_obj.exam,
                overall_gpa__isnull=False,
            ).select_related('student').order_by(
                db_models.F('class_rank').asc(nulls_last=True),
                db_models.F('overall_gpa').desc(nulls_last=True),
            )[:5]
            for idx, r in enumerate(results_qs):
                top_students.append({
                    'rank': r.class_rank if r.class_rank else idx + 1,
                    'student': r.student,
                    'gpa': float(r.overall_gpa),
                    'grade': r.final_grade,
                    'is_pass': r.is_pass,
                })

        # --- At-risk students (GPA < 2.0 or is_pass=False) ---
        at_risk_qs = StudentResult.objects.filter(
            student__class_obj=cls_obj,
        )
        if latest_ec_obj:
            at_risk_qs = at_risk_qs.filter(exam=latest_ec_obj.exam)
        elif active_session:
            at_risk_qs = at_risk_qs.filter(session=active_session)
        at_risk = at_risk_qs.filter(
            db_models.Q(is_pass=False) | db_models.Q(overall_gpa__lt=2.0)
        ).select_related('student').order_by(
            db_models.F('overall_gpa').asc(nulls_last=True)
        )[:10]
        at_risk_list = [{
            'student': r.student,
            'gpa': float(r.overall_gpa) if r.overall_gpa is not None else None,
            'grade': r.final_grade,
            'failed_subjects': r.failed_subjects,
        } for r in at_risk]

        # --- Attendance summary ---
        # Get the latest exam in this class
        latest_entries = MarkEntry.objects.filter(
            student__class_obj=cls_obj,
            session=active_session,
            total_days__isnull=False,
            total_days__gt=0,
        ).values('student_id', 'present_days', 'total_days').distinct()

        attendance_data = {}
        for entry in latest_entries:
            sid = entry['student_id']
            if sid not in attendance_data:
                attendance_data[sid] = entry

        total_days_val = None
        attendance_list = []
        for sid, entry in attendance_data.items():
            pd = entry['present_days'] or 0
            td = entry['total_days'] or 0
            if td > 0:
                if total_days_val is None:
                    total_days_val = td
                pct = round(pd / td * 100, 1)
                attendance_list.append({'student_id': sid, 'present': pd, 'total': td, 'pct': pct})

        # map student_id -> name
        student_map = {s.id: s for s in cls_students}
        for a in attendance_list:
            a['student'] = student_map.get(a['student_id'])

        low_attendance = sorted([a for a in attendance_list if a['pct'] < 75], key=lambda x: x['pct'])
        avg_attendance = round(sum(a['pct'] for a in attendance_list) / len(attendance_list), 1) if attendance_list else None

        class_analytics.append({
            'class_obj': cls_obj,
            'subject_averages': subject_averages,
            'top_students': top_students,
            'at_risk': at_risk_list,
            'low_attendance': low_attendance,
            'avg_attendance': avg_attendance,
            'total_days': total_days_val,
            'attendance_tracked': bool(attendance_list),
        })

    # ── SUBJECT AVERAGES: computed for ALL teachers with assigned subjects ──
    is_class_teacher_dashboard = bool(my_classes_info)
    subject_teacher_averages = []
    if my_subject_list:
        from apps.marks.models import MarkEntry as ME
        from django.db.models import Avg as DAvg
        for item in my_subject_list:
            sub = item['subject']
            cls_obj = item['class_obj']
            avg_qs = ME.objects.filter(
                subject=sub,
                student__class_obj=cls_obj,
                session=active_session,
            )
            if recent_exam:
                avg_qs = avg_qs.filter(exam=recent_exam)
                
            avg = avg_qs.aggregate(th_avg=DAvg('theory_obtained'), in_avg=DAvg('internal_obtained'))
            th_avg = float(avg['th_avg'] or 0)
            in_avg = float(avg['in_avg'] or 0)
            total_avg = th_avg + in_avg
            full = float((sub.theory_full_marks or 0) + (sub.practical_full_marks or 0))
            pct = round((total_avg / full * 100), 1) if full > 0 else 0
            subject_teacher_averages.append({
                'label': f"{cls_obj.full_name} — {sub.name}",
                'avg': round(total_avg, 1),
                'full': full,
                'pct': pct,
            })

    return render(request, 'teachers/teacher_portal/dashboard.html', {
        'teacher': teacher,
        'my_classes_info': my_classes_info,
        'my_subjects_info': my_subjects_info,
        'my_subject_list': my_subject_list,
        'class_analytics': class_analytics,
        'is_class_teacher_dashboard': is_class_teacher_dashboard,
        'subject_teacher_averages': subject_teacher_averages,
        'total_classes': total_classes,
        'total_students': total_students,
        'total_subjects': total_subjects,
        'primary_subject': primary_subject,
        'pending_marks': pending_marks,
        'published_percentage': published_percentage,
        'avg_gpa': avg_gpa,
        'grade_labels_json': grade_labels_json,
        'grade_data_json': grade_data_json,
        'top_performers': top_performers,
        'recent_exam_name': recent_exam.name if recent_exam else None,
    })


@login_required
def teacher_spreadsheet_edit(request):
    school = request.user.school
    active_session = school.get_active_session() if school else None
    
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('teacher_list')
        
    teachers_qs = Teacher.objects.filter(school=school)
    if active_session:
        teachers_qs = teachers_qs.filter(sessions=active_session)
        
    def get_sort_key(t):
        try:
            val = float(t.sn) if t.sn else float('inf')
        except ValueError:
            import re
            nums = re.findall(r'\d+', t.sn)
            val = float(nums[0]) if nums else float('inf')
        return (val, t.name)
        
    teachers = sorted(teachers_qs, key=get_sort_key)

    return render(request, 'teachers/spreadsheet_edit.html', {
        'teachers': teachers,
    })


@login_required
def teacher_inline_edit(request, teacher_id):
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)
    
    school = request.user.school
    teacher = get_object_or_404(Teacher, pk=teacher_id, school=school)
    
    if request.user.role not in [request.user.Role.SUPER_ADMIN, request.user.Role.SCHOOL_ADMIN]:
        return JsonResponse({'status': 'error', 'message': 'Access denied'}, status=403)
        
    field_name = request.POST.get('name')
    value = request.POST.get('value')
    
    allowed_fields = ['sn', 'name', 'contact_number', 'email']
    if field_name not in allowed_fields:
        return JsonResponse({'status': 'error', 'message': 'Invalid field'}, status=400)
        
    try:
        if field_name == 'email' and value:
            if Teacher.objects.filter(email__iexact=value).exclude(pk=teacher_id).exists():
                return JsonResponse({'status': 'error', 'message': 'Email is already in use.'}, status=400)
        
        if field_name == 'email' and not value:
            value = None
            
        setattr(teacher, field_name, value)
        teacher.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
def teacher_subject_map(request, pk):
    school = request.user.school
    if request.user.role not in [User.Role.SUPER_ADMIN, User.Role.SCHOOL_ADMIN]:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
        
    teacher = get_object_or_404(Teacher, pk=pk, school=school)
    
    if request.method == 'POST':
        subject_ids = request.POST.getlist('subject_ids')
        
        with transaction.atomic():
            TeacherSubject.objects.filter(teacher=teacher).delete()
            new_mappings = []
            for sid in subject_ids:
                if sid.isdigit():
                    new_mappings.append(TeacherSubject(teacher=teacher, subject_id=int(sid)))
            TeacherSubject.objects.bulk_create(new_mappings)
            
        messages.success(request, f'Subjects successfully mapped to {teacher.name}.')
        return redirect('teacher_detail', pk=teacher.pk)

    active_session = school.get_active_session()
    
    # Get all classes in active session with subjects
    classes = Class.objects.filter(school=school, session=active_session).prefetch_related('subjects')
    
    # Pre-calculate what is mapped
    all_mappings = TeacherSubject.objects.filter(teacher__school=school).select_related('teacher')
    mapping_dict = {tm.subject_id: tm.teacher for tm in all_mappings}
    current_mapped_subject_ids = set(TeacherSubject.objects.filter(teacher=teacher).values_list('subject_id', flat=True))
    
    for cls in classes:
        for sub in cls.subjects.all():
            if sub.id in current_mapped_subject_ids:
                sub.is_mapped_to_current = True
                sub.other_teacher_name = None
            else:
                sub.is_mapped_to_current = False
                other_teacher = mapping_dict.get(sub.id)
                sub.other_teacher_name = other_teacher.name if other_teacher else None

    return render(request, 'teachers/subject_map.html', {
        'teacher': teacher,
        'classes': classes
    })
